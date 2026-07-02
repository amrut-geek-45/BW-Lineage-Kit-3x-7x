import streamlit as st
import pandas as pd
import os
import glob
import re
import io
import zipfile
from collections import Counter, defaultdict

from analyzer import analyze_code
from transformation_code import process_transformation

# =====================================================
# PAGE CONFIG
# =====================================================

st.set_page_config(
    page_title="BW Transformation Analyzer",
    page_icon="⚙️",
    layout="wide"
)

# =====================================================
# PASSWORD GATE
# Password stored in .streamlit/secrets.toml:
#   APP_PASSWORD = "your-team-password"
# =====================================================

def check_password():
    """Returns True if user entered the correct password."""
    if st.session_state.get("authenticated"):
        return True

    # Centre the login card
    _, col, _ = st.columns([1, 1.2, 1])
    with col:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown(
            "<div style='text-align:center;font-size:32px;margin-bottom:4px'>⚙️</div>"
            "<div style='text-align:center;font-size:20px;font-weight:600;"
            "color:#111827;margin-bottom:4px'>BW Migration Toolkit</div>"
            "<div style='text-align:center;font-size:13px;color:#64748b;"
            "margin-bottom:24px'>Enter your team password to continue</div>",
            unsafe_allow_html=True
        )
        pwd = st.text_input("Password", type="password",
                            placeholder="Enter password…", key="pwd_input",
                            label_visibility="collapsed")
        if st.button("Sign in", type="primary", use_container_width=True):
            try:
                correct = st.secrets.get("APP_PASSWORD", "")
            except:
                correct = ""
            if correct and pwd == correct:
                st.session_state.authenticated = True
                st.rerun()
            elif not correct:
                st.warning("⚠️ APP_PASSWORD not set in secrets.toml")
            else:
                st.error("❌ Incorrect password — please try again")
        st.markdown(
            "<div style='text-align:center;font-size:11px;color:#94a3b8;"
            "margin-top:16px'>Contact your admin for access</div>",
            unsafe_allow_html=True
        )
    st.stop()

check_password()

# =====================================================
# CSS
# =====================================================

st.markdown("""
<style>
/* ══ FORCE EVERYTHING DARK TEXT ON WHITE ══ */
.stApp, .stApp > div, .main, .block-container,
[data-testid="stAppViewContainer"],
[data-testid="stAppViewBlockContainer"] {
    background: #ffffff !important;
    color: #111827 !important;
}
*, *::before, *::after { color: #111827; }
[data-testid="stSidebar"],
[data-testid="stSidebar"] > div,
[data-testid="stSidebar"] section {
    background: #f5f7fa !important;
    border-right: 1px solid #e2e8f0 !important;
}
[data-testid="stSidebar"] * { color: #111827 !important; }
.hero-title { font-size:40px; font-weight:700; color:#111827 !important; margin-bottom:4px }
.hero-sub   { font-size:16px; color:#374151 !important; margin-bottom:1.5rem }
h1,h2,h3,h4,h5,h6 { color:#111827 !important; font-weight:700 }
p, span, div, li, td, th, label, caption,
.stMarkdown p, .stMarkdown span,
.stMarkdown li, .stMarkdown div { color:#111827 !important }
.stCaption, small, [data-testid="stCaptionContainer"],
[data-testid="stCaptionContainer"] p { color:#374151 !important }
.stTabs [data-baseweb="tab-list"] {
    background:#f1f5f9 !important; border-radius:8px !important;
    padding:4px !important; gap:2px !important; border:none !important;
}
.stTabs [data-baseweb="tab"] {
    background:transparent !important; color:#374151 !important;
    border-radius:6px !important; font-weight:500 !important;
    font-size:13px !important; padding:6px 14px !important; border:none !important;
}
.stTabs [data-baseweb="tab"] p,
.stTabs [data-baseweb="tab"] span,
.stTabs [data-baseweb="tab"] div { color:#374151 !important }
.stTabs [data-baseweb="tab"]:hover { background:#e2e8f0 !important; color:#111827 !important; }
.stTabs [aria-selected="true"],
.stTabs [aria-selected="true"] p,
.stTabs [aria-selected="true"] span,
.stTabs [aria-selected="true"] div {
    background:#ffffff !important; color:#2563eb !important;
    font-weight:600 !important; box-shadow:0 1px 4px rgba(0,0,0,.12) !important;
}
.stTabs [data-baseweb="tab-highlight"],
.stTabs [data-baseweb="tab-border"] { background:transparent !important; height:0 !important; }
.stTabs [data-baseweb="tab-panel"] { background:#ffffff !important; padding-top:12px !important; }
.stTextInput > label, .stTextArea > label,
.stSelectbox > label, .stSlider > label { color:#111827 !important; font-weight:500 }
.stTextInput input, .stTextArea textarea {
    background:#ffffff !important; color:#111827 !important;
    border:1px solid #d1d5db !important; border-radius:6px !important;
}
.stTextInput input::placeholder,
.stTextArea textarea::placeholder { color:#6b7280 !important }
[data-testid="stFileUploader"] {
    background:#f8fafc !important; border:1.5px dashed #cbd5e1 !important; border-radius:8px !important;
}
[data-testid="stFileUploader"] *,
[data-testid="stFileUploaderDropzoneInstructions"] * { color:#111827 !important }
[data-testid="stBaseButton-secondary"] {
    background:#ffffff !important; color:#111827 !important; border:1px solid #d1d5db !important;
}
.stButton > button {
    background:#f1f5f9 !important; color:#111827 !important;
    border:1px solid #d1d5db !important; border-radius:6px !important; font-weight:500 !important;
}
.stButton > button:hover { background:#e2e8f0 !important; border-color:#94a3b8 !important; }
.stButton > button[kind="primary"], button[kind="primary"] {
    background:#2563eb !important; color:#ffffff !important; border:none !important;
}
.stButton > button[kind="primary"]:hover { background:#1d4ed8 !important }
div[data-testid="stDownloadButton"] button {
    background:#2563eb !important; color:#ffffff !important;
    border:none !important; border-radius:6px !important;
    font-weight:600 !important; padding:8px 20px !important;
}
div[data-testid="stDownloadButton"] button:hover { background:#1d4ed8 !important }
[data-testid="stToggleLabel"],
.stToggle label, .stToggle p, .stToggle span,
[data-testid="stWidgetLabel"] p,
[data-testid="stWidgetLabel"] span { color:#111827 !important; font-weight:500 !important }
.stRadio label, .stRadio p, .stRadio span { color:#111827 !important }
div[data-testid="metric-container"] {
    background:#f8fafc !important; border:1px solid #e2e8f0 !important;
    border-radius:10px !important; padding:12px 16px !important;
}
div[data-testid="stMetricLabel"],
div[data-testid="stMetricLabel"] p,
div[data-testid="stMetricLabel"] span { color:#374151 !important }
div[data-testid="stMetricValue"],
div[data-testid="stMetricValue"] p { color:#111827 !important; font-weight:700 !important }
.streamlit-expanderHeader,
.streamlit-expanderHeader * {
    background:#f8fafc !important; color:#111827 !important;
    border:1px solid #e2e8f0 !important; border-radius:6px !important;
}
.streamlit-expanderContent {
    background:#ffffff !important; border:1px solid #e2e8f0 !important; border-top:none !important;
}
.streamlit-expanderContent * { color:#111827 !important }
.stSuccess { background:#f0fdf4 !important }
.stSuccess *, .stSuccess p { color:#14532d !important }
.stWarning { background:#fffbeb !important }
.stWarning *, .stWarning p { color:#92400e !important }
.stError { background:#fff1f2 !important }
.stError *, .stError p { color:#9f1239 !important }
.stInfo { background:#eff8ff !important }
.stInfo *, .stInfo p { color:#1e3a8a !important }
.stDataFrame { border:1px solid #e2e8f0 !important; border-radius:8px !important }
.stDataFrame * { color:#111827 !important }
.stDataFrame thead th {
    background:#f1f5f9 !important; color:#111827 !important; font-weight:600 !important;
}
[data-testid="stChatInput"] {
    background:#f8fafc !important; border:1px solid #e2e8f0 !important; border-radius:8px !important;
}
[data-testid="stChatInput"] textarea,
[data-testid="stChatInput"] * { color:#111827 !important }
[data-testid="stChatInput"] textarea::placeholder { color:#6b7280 !important }
.stProgress > div > div { background:#2563eb !important }
hr { border-color:#e2e8f0 !important }
.metric-card {
    background:#f8fafc; padding:20px; border-radius:12px;
    border:1px solid #e2e8f0; text-align:center;
}
.metric-card h3 { font-size:12px; color:#374151; margin-bottom:6px; text-transform:uppercase; letter-spacing:.06em }
.metric-card h2 { font-size:28px; font-weight:700; margin:0; color:#111827 }
.cx-LOW    { color:#16a34a !important }
.cx-MEDIUM { color:#d97706 !important }
.cx-HIGH   { color:#dc2626 !important }
.ai-box, .ai-sum-box {
    background:#f5f3ff; border-left:4px solid #7c3aed;
    padding:14px 18px; border-radius:6px; font-size:14px; line-height:1.7; margin:8px 0;
}
.ai-box *, .ai-sum-box * { color:#111827 !important }
.ztable-card {
    background:#f8fafc; border:1px solid #e2e8f0; border-radius:8px;
    padding:10px 14px; font-size:13px; color:#111827; margin:4px 0;
}
.ln-wrap { font-family:monospace; margin:2px 0 }
.ln-row {
    display:flex; align-items:center; gap:6px; padding:6px 12px;
    border-radius:6px; margin:2px 0; font-size:13px;
    white-space:normal; overflow:visible; flex-wrap:wrap;
}
.ln-row * { color:#111827 !important }
.ln-query       { background:#f0effe; border-left:4px solid #7c3aed }
.ln-provider    { background:#eff8ff; border-left:4px solid #2563eb }
.ln-mp          { background:#eff8ff; border-left:4px solid #1d4ed8 }
.ln-cube        { background:#f0fdf4; border-left:4px solid #16a34a }
.ln-adso        { background:#fffbeb; border-left:4px solid #d97706 }
.ln-trfn        { background:#faf5ff; border-left:4px solid #9333ea }
.ln-dtp         { background:#f0fdfa; border-left:4px solid #0891b2 }
.ln-datasource  { background:#fff1f2; border-left:4px solid #e11d48 }
.ln-infosource  { background:#eff8ff; border-left:4px solid #2563eb }
.ln-infopackage { background:#f0effe; border-left:4px solid #7c3aed }
.ln-srcsys      { background:#fefce8; border-left:4px solid #ca8a04 }
.ln-default     { background:#f8fafc; border-left:4px solid #94a3b8 }
.ln-connector   { opacity:.4; margin-right:2px; user-select:none; flex-shrink:0 }
.ln-type        { font-size:10px; color:#374151 !important; font-weight:700; text-transform:uppercase; letter-spacing:.05em; flex-shrink:0; white-space:nowrap }
.ln-name        { font-size:12px; color:#111827 !important; background:rgba(0,0,0,.06); padding:1px 6px; border-radius:3px; word-break:break-all; white-space:normal }
.ln-legend      { display:flex; flex-wrap:wrap; gap:6px; margin-top:10px }
.ln-leg-item    { display:flex; align-items:center; gap:5px; padding:3px 10px; border-radius:20px; font-size:11px }
.ln-leg-item *  { color:#111827 !important }
/* downstream special */
.ln-downstream  { background:#fefce8; border-left:4px solid #ca8a04 }
.ln-section-hdr { font-size:12px; font-weight:700; color:#374151; text-transform:uppercase;
                  letter-spacing:.07em; padding:6px 0 3px 0; border-bottom:1px solid #e2e8f0;
                  margin:10px 0 6px 0; }
.risk-CRITICAL  { background:#fff1f2; border-left:5px solid #e11d48; padding:12px 16px; border-radius:6px; font-size:13px; margin:6px 0 }
.risk-CRITICAL * { color:#9f1239 !important }
.risk-HIGH      { background:#fffbeb; border-left:5px solid #d97706; padding:12px 16px; border-radius:6px; font-size:13px; margin:6px 0 }
.risk-HIGH *    { color:#92400e !important }
.risk-MEDIUM    { background:#f0fdf4; border-left:5px solid #16a34a; padding:12px 16px; border-radius:6px; font-size:13px; margin:6px 0 }
.risk-MEDIUM *  { color:#14532d !important }
.risk-LOW       { background:#eff8ff; border-left:5px solid #2563eb; padding:12px 16px; border-radius:6px; font-size:13px; margin:6px 0 }
.risk-LOW *     { color:#1e3a8a !important }
.risk-banner-CRITICAL { background:#fff1f2; border:1px solid #e11d48; padding:14px 18px; border-radius:8px; margin:8px 0 }
.risk-banner-CRITICAL * { color:#9f1239 !important }
.risk-banner-HIGH { background:#fffbeb; border:1px solid #d97706; padding:14px 18px; border-radius:8px; margin:8px 0 }
.risk-banner-HIGH * { color:#92400e !important }
.risk-banner-MEDIUM { background:#f0fdf4; border:1px solid #16a34a; padding:14px 18px; border-radius:8px; margin:8px 0 }
.risk-banner-MEDIUM * { color:#14532d !important }
.risk-banner-LOW { background:#eff8ff; border:1px solid #2563eb; padding:14px 18px; border-radius:8px; margin:8px 0 }
.risk-banner-LOW * { color:#1e3a8a !important }
.impact-hit  { background:#fffbeb; border-left:4px solid #d97706; padding:9px 12px; border-radius:5px; font-size:13px; margin:4px 0 }
.impact-hit * { color:#92400e !important }
.impact-none { background:#f0fdf4; border-left:4px solid #16a34a; padding:9px 12px; border-radius:5px; font-size:13px }
.impact-none *{ color:#14532d !important }
.dead-card   { background:#fff1f2; border-left:4px solid #e11d48; padding:8px 12px; border-radius:5px; font-size:12px; margin:3px 0 }
.dead-card * { color:#9f1239 !important }
.compare-box    { background:#f8fafc; border:1px solid #e2e8f0; border-radius:8px; padding:14px; font-size:13px; margin:4px 0 }
.compare-box *  { color:#111827 !important }
.compare-shared { background:#f0fdf4; border-left:3px solid #16a34a; padding:5px 10px; border-radius:4px; font-size:12px; margin:2px 0 }
.compare-shared *{ color:#14532d !important }
.compare-only   { background:#eff8ff; border-left:3px solid #2563eb; padding:5px 10px; border-radius:4px; font-size:12px; margin:2px 0 }
.compare-only * { color:#1e3a8a !important }
.chat-user { background:#eff8ff; border-radius:12px 12px 2px 12px; padding:10px 14px; margin:5px 0; font-size:13px }
.chat-user *{ color:#111827 !important }
.chat-ai   { background:#f5f3ff; border:1px solid #e9d5ff; border-radius:2px 12px 12px 12px; padding:10px 14px; margin:5px 0; font-size:13px; line-height:1.6 }
.chat-ai * { color:#111827 !important }
.tool-card   { background:#f8fafc; border:1px solid #e2e8f0; border-radius:10px; padding:16px; margin-bottom:14px }
.tool-card * { color:#111827 !important }
.upload-card   { background:#f8fafc; border:0.5px solid #e2e8f0; border-radius:8px; padding:14px; margin-bottom:10px }
.upload-card * { color:#111827 !important }
</style>
""", unsafe_allow_html=True)

# =====================================================
# METADATA FILE DEFINITIONS
# =====================================================

META_FILES = {
    "query_df": {
        "label":    "📋 Query File",
        "desc":     "QUERYID · QUERYNAME · INFOPROVIDER",
        "required": True,
    },
    "trfn_df": {
        "label":    "🔁 Transformation File",
        "desc":     "TRANID · SOURCENAME · TARGETNAME",
        "required": True,
    },
    "ip_df": {
        "label":    "📬 InfoPackage File",
        "desc":     "LOGDPID · OLTPSOURCE",
        "required": True,
    },
    "cube_df": {
        "label":    "🧊 InfoCube File",
        "desc":     "INFOCUBE column",
        "required": False,
    },
    "adso_df": {
        "label":    "🗄️ ADSO File",
        "desc":     "ADSO column",
        "required": False,
    },
    "mp_df": {
        "label":    "🔀 MultiProvider File",
        "desc":     "MULTIPROVIDER · PARTPROVIDER",
        "required": False,
    },
    "upd_df": {
        "label":    "🔄 Update Rules File",
        "desc":     "UPDID · INFOCUBE · ISOURCE",
        "required": False,
    },
    "ds_df": {
        "label":    "📡 DataSource File",
        "desc":     "DATASOURCE · LOGSYS · TXTSH",
        "required": False,
    },
    "ext_df": {
        "label":    "🔌 Extractor File",
        "desc":     "Communication Structure · CommunicationsStructure techn.",
        "required": False,
    },
}

# =====================================================
# METADATA LOADERS  (upload-based — no server folder)
# =====================================================

def _read_uploaded_df(uf):
    """
    Read a Streamlit UploadedFile into a DataFrame.

    Two strategies balanced against each other:

    A) PANDAS (used when ALL rows have the same field count):
       Correct column alignment even for wide files (ip_df has 24 cols).
       Drops rows where description columns contain the separator — bad for query files.

    B) SMART PARSER (used when rows have variable field counts):
       Never drops rows — merges overflow fields back into last column.
       Can cause column shift for wide files if data rows have fewer fields than header.

    Solution: try pandas first. If it drops rows vs the smart parser count,
    use smart parser result instead. Best of both worlds.
    """
    if uf is None:
        return None
    try:
        name = uf.name.lower()
        if name.endswith((".xls", ".xlsx")):
            uf.seek(0)
            return pd.read_excel(uf)

        # ── Read raw bytes with encoding fallback
        uf.seek(0)
        raw_bytes = uf.read()
        text = None
        for enc in ["utf-8", "utf-8-sig", "cp1252", "latin1", "ISO-8859-1"]:
            try:
                text = raw_bytes.decode(enc); break
            except Exception:
                pass
        if text is None:
            text = raw_bytes.decode("cp1252", errors="replace")

        lines = [l for l in text.splitlines() if l.strip()]
        if not lines:
            return None

        # ── Detect separator: most occurrences in header wins
        header_line = lines[0].lstrip("﻿")
        best_sep    = "	"
        best_count  = 0
        for candidate in ["	", ",", ";", "|"]:
            c = header_line.count(candidate)
            if c > best_count:
                best_count = c
                best_sep   = candidate

        header = [c.strip().strip('"') for c in header_line.split(best_sep)]
        n_cols = len(header)
        if n_cols < 1:
            return None

        data_lines = [l for l in lines[1:] if l.strip()]
        total_data = len(data_lines)

        # ── Strategy A: pandas (correct alignment for wide/clean files)
        df_pandas = None
        for enc in ["utf-8", "utf-8-sig", "cp1252", "latin1", "ISO-8859-1"]:
            try:
                uf.seek(0)
                df_try = pd.read_csv(uf, encoding=enc, sep=best_sep,
                                     engine="python", on_bad_lines="skip")
                if df_try is not None and len(df_try.columns) == n_cols:
                    df_pandas = df_try
                    break
            except Exception:
                pass

        # ── Strategy B: smart parser (no rows dropped, overflow merged into last col)
        records = []
        for line in data_lines:
            parts = line.split(best_sep)
            if len(parts) >= n_cols:
                row = parts[:n_cols-1] + [best_sep.join(parts[n_cols-1:])]
            else:
                row = parts + [""] * (n_cols - len(parts))
            records.append([f.strip().strip('"') for f in row])
        df_smart = pd.DataFrame(records, columns=header) if records else None

        # ── Pick the best result:
        # - If pandas got same row count as smart parser → use pandas (correct alignment)
        # - If pandas dropped rows (< smart count) → use smart parser (no lost rows)
        # - If pandas failed entirely → use smart parser
        if df_pandas is not None and df_smart is not None:
            if len(df_pandas) >= len(df_smart):
                return df_pandas   # pandas has correct alignment, no rows lost
            else:
                return df_smart    # smart parser recovers dropped rows
        elif df_pandas is not None:
            return df_pandas
        elif df_smart is not None:
            return df_smart

    except Exception:
        pass
    return None


def _filter_active(df):
    """Keep only OBJVERS = A rows. Works whether column exists or not."""
    if df is None:
        return None
    df = df.astype(str)
    if "OBJVERS" in df.columns:
        df = df[df["OBJVERS"].str.strip().str.upper() == "A"].copy()
    return df


def _detect_file_key(df):
    """
    Auto-detect which metadata file type a DataFrame is
    by checking for key columns unique to each file type.
    Returns the META_FILES key or None.
    """
    cols = set(df.columns.str.upper().str.strip())

    # Order matters — check most specific first
    if "QUERYID"                  in cols and "INFOPROVIDER"  in cols:                              return "query_df"
    if "TRANID"                   in cols and "SOURCENAME"    in cols and "TARGETNAME" in cols:     return "trfn_df"
    if "LOGDPID"                  in cols and "OLTPSOURCE"    in cols:                              return "ip_df"
    if "MULTIPROVIDER"            in cols and "PARTPROVIDER"  in cols:                              return "mp_df"
    if "UPDID"                    in cols and "INFOCUBE"       in cols and "ISOURCE"   in cols:     return "upd_df"
    if "DATASOURCE"               in cols and "LOGSYS"        in cols and "TXTSH"      in cols:     return "ds_df"
    if "COMMUNICATION STRUCTURE"  in cols and "STAT"          in cols:                              return "ext_df"
    if "INFOCUBE"                 in cols:                                                           return "cube_df"
    if "ADSO"                     in cols:                                                           return "adso_df"
    return None


def parse_zip_upload(zip_file):
    """
    Extract all CSV/Excel files from an uploaded ZIP,
    auto-detect each file's type, return {key: UploadedFile-like object}.
    Returns (uploads_dict, messages_list).
    """
    import zipfile as _zipfile, io as _io

    uploads  = {}
    messages = []

    zip_file.seek(0)
    try:
        zf = _zipfile.ZipFile(zip_file)
    except Exception as e:
        return {}, [f"❌ Could not open ZIP file: {e}"]

    for name in zf.namelist():
        # Skip folders and hidden files
        if name.endswith("/") or name.startswith("__") or name.startswith("."):
            continue
        lower = name.lower()
        if not any(lower.endswith(ext) for ext in (".csv",".txt",".tsv",".xls",".xlsx")):
            continue

        try:
            raw = zf.read(name)
            buf = _io.BytesIO(raw)
            buf.name = name.split("/")[-1]  # just filename, no path
            buf.size = len(raw)
            buf.seek(0)

            df = _read_uploaded_df(buf)
            if df is None or len(df) == 0:
                messages.append(f"⚠️ {buf.name} — could not read or empty")
                continue

            key = _detect_file_key(df)
            if key is None:
                messages.append(f"⚠️ {buf.name} — columns not recognised, skipped")
                continue

            if key in uploads:
                messages.append(f"⚠️ {buf.name} — duplicate {key}, keeping first")
                continue

            buf.seek(0)
            uploads[key] = buf
            messages.append(f"✅ {buf.name} → {META_FILES[key]['label']}")

        except Exception as e:
            messages.append(f"❌ {name} — error: {e}")

    zf.close()
    return uploads, messages


def prepare_dfs(uploads):
    """
    Takes {key: UploadedFile} dict.
    Returns fully prepared dfs dict ready for all lineage functions.
    """
    dfs = {}
    # InfoPackages use different OBJVERS values — don't filter them
    for key, uf in uploads.items():
        raw = _read_uploaded_df(uf)
        df  = _filter_active(raw)  # keeps OBJVERS=A for all files including ip_df
        if df is not None and len(df) > 0:
            dfs[key] = df

    # Clean client suffix (GDVCLNTxxx) from object names in ALL relevant files
    CLIENT_COLS = {
        "trfn_df": ["SOURCENAME","TARGETNAME"],
        "ip_df":   ["OLTPSOURCE","SOURCE","LOGDPID","LOGDPID_N"],
    }
    # Two-pass client suffix removal:
    # Pass 1: remove " QMDCLNT100" style (space-separated, exactly 3 prefix chars)
    # Pass 2: remove "GDVCLNT100" style (concatenated, 3 prefix chars + CLNT + digits)
    # Covers: GDVCLNT, QMDCLNT, BWPCLNT, DEVCLNT, PRDCLNT, PROCLNT etc.
    import re as _re
    _P1 = r"\s+[A-Z]{3}CLNT\d+"
    _P2 = r"(?<=[A-Z0-9_])[A-Z]{3}CLNT\d+"
    def _strip_client(val):
        """Remove SAP client suffixes: GDVCLNT100, QMDCLNT100, space-separated or concatenated."""
        v = _re.sub(_P1, "", str(val))
        v = _re.sub(_P2, "", v)
        return v.strip()

    for df_key, cols in CLIENT_COLS.items():
        if dfs.get(df_key) is not None:
            for col in cols:
                if col in dfs[df_key].columns:
                    dfs[df_key][col] = dfs[df_key][col].astype(str).apply(_strip_client)

    # O(1) lookup sets for object type detection
    dfs["_mp_set"]   = set(dfs["mp_df"]["MULTIPROVIDER"].str.upper())  if dfs.get("mp_df")   is not None else set()
    dfs["_cube_set"] = set(dfs["cube_df"]["INFOCUBE"].str.upper())     if dfs.get("cube_df") is not None else set()
    dfs["_adso_set"] = set(dfs["adso_df"]["ADSO"].str.upper())         if dfs.get("adso_df") is not None else set()

    # ── Index dictionaries — O(1) lookups replacing O(n) DataFrame scans ──
    # Built once at load time; used throughout build_lineage + build_downstream

    # trfn_by_target[TARGET_UPPER] = list of row dicts {TRANID, SOURCENAME}
    idx = {}
    if dfs.get("trfn_df") is not None and "TARGETNAME" in dfs["trfn_df"].columns:
        for _, r in dfs["trfn_df"].iterrows():
            k = str(r["TARGETNAME"]).upper().strip()
            idx.setdefault(k, []).append({
                "TRANID":      str(r.get("TRANID","")).strip(),
                "SOURCENAME":  str(r.get("SOURCENAME","")).strip(),
            })
    dfs["_idx_trfn_by_target"] = idx

    # trfn_by_source[SOURCE_UPPER] = list of row dicts {TRANID, TARGETNAME}
    idx2 = {}
    if dfs.get("trfn_df") is not None and "SOURCENAME" in dfs["trfn_df"].columns:
        for _, r in dfs["trfn_df"].iterrows():
            k = str(r["SOURCENAME"]).upper().strip()
            idx2.setdefault(k, []).append({
                "TRANID":     str(r.get("TRANID","")).strip(),
                "TARGETNAME": str(r.get("TARGETNAME","")).strip(),
            })
    dfs["_idx_trfn_by_source"] = idx2

    # ip_by_oltpsource[OLTPSOURCE_UPPER] = list of {LOGDPID}
    # Handles: OLTPSOURCE or SOURCE column; LOGDPID or LOGDPID_N column
    idx4 = {}
    ip_df_raw = dfs.get("ip_df")
    if ip_df_raw is not None:
        ip_cols = ip_df_raw.columns.tolist()
        # Determine which column holds the DataSource name
        src_col  = "OLTPSOURCE" if "OLTPSOURCE" in ip_cols else ("SOURCE" if "SOURCE" in ip_cols else None)
        # Determine which column holds the InfoPackage ID
        # Prefer LOGDPID_N (unique ID), fall back to LOGDPID
        ipid_col = "LOGDPID_N" if "LOGDPID_N" in ip_cols else ("LOGDPID" if "LOGDPID" in ip_cols else None)
        if src_col and ipid_col:
            for _, r in ip_df_raw.iterrows():
                k     = str(r[src_col]).upper().strip()
                ip_id = str(r.get(ipid_col,"")).strip()
                # Also store under LOGDPID if different from LOGDPID_N
                if ip_id.lower() not in ("nan","none","") and k.lower() not in ("nan","none",""):
                    idx4.setdefault(k, [])
                    if ip_id not in idx4[k]:
                        idx4[k].append(ip_id)
            # If we used LOGDPID_N, also index by LOGDPID as secondary ID
            if ipid_col == "LOGDPID_N" and "LOGDPID" in ip_cols:
                for _, r in ip_df_raw.iterrows():
                    k      = str(r[src_col]).upper().strip()
                    ip_id2 = str(r.get("LOGDPID","")).strip()
                    if ip_id2.lower() not in ("nan","none","") and k.lower() not in ("nan","none",""):
                        idx4.setdefault(k, [])
                        if ip_id2 not in idx4[k]:
                            idx4[k].append(ip_id2)
    dfs["_idx_ip"] = idx4
    dfs["_ip_src_col"]  = src_col  if ip_df_raw is not None else None
    dfs["_ip_id_col"]   = ipid_col if ip_df_raw is not None else None

    # mp_parts[MULTIPROVIDER_UPPER] = list of PARTPROVIDER strings
    idx5 = {}
    if dfs.get("mp_df") is not None and "MULTIPROVIDER" in dfs["mp_df"].columns and "PARTPROVIDER" in dfs["mp_df"].columns:
        for _, r in dfs["mp_df"].iterrows():
            k    = str(r["MULTIPROVIDER"]).upper().strip()
            part = str(r.get("PARTPROVIDER","")).strip()
            if part.lower() not in ("nan","none",""):
                idx5.setdefault(k, []).append(part)
    dfs["_idx_mp_parts"] = idx5

    # upd_by_cube[INFOCUBE_UPPER] = list of {UPDID, ISOURCE, OBJSTAT}
    # Built from Update Rules file (RSUPDINFO) — maps cubes to their update programs and ISOURCEs
    idx7 = {}
    if dfs.get("upd_df") is not None:
        udf = dfs["upd_df"]
        for col in ["UPDID","INFOCUBE","ISOURCE","OBJSTAT"]:
            if col in udf.columns:
                udf[col] = udf[col].astype(str).str.strip()
        if "INFOCUBE" in udf.columns and "ISOURCE" in udf.columns:
            for _, r in udf.iterrows():
                k  = str(r.get("INFOCUBE","")).upper().strip()
                st = str(r.get("OBJSTAT","ACT")).strip()
                if k.lower() in ("nan","none",""):
                    continue
                idx7.setdefault(k, []).append({
                    "updid":   str(r.get("UPDID","")).strip(),
                    "isource": str(r.get("ISOURCE","")).strip(),
                    "objstat": st,
                })
    dfs["_idx_upd_by_cube"] = idx7

    # query_by_provider[INFOPROVIDER_UPPER] = list of "QUERYID – description" strings
    # Also builds a reverse cube→MP index so downstream chains on cubes can find MP-level queries
    idx6 = {}
    if dfs.get("query_df") is not None and "INFOPROVIDER" in dfs["query_df"].columns:
        qdf = dfs["query_df"]
        for _, r in qdf.iterrows():
            k    = str(r["INFOPROVIDER"]).upper().strip()
            qid  = str(r.get("QUERYID","")).strip()
            desc = str(r.get("SHORTTEXT", r.get("LONGTEXT",""))).strip()
            if qid.lower() in ("nan","none",""): continue
            label = f"{qid} – {desc}" if desc and desc.lower() not in ("nan","none","") else qid
            idx6.setdefault(k, []).append(label)
    dfs["_idx_queries"] = idx6

    # cube_to_mp[CUBE_UPPER] = list of MULTIPROVIDER names that contain this cube
    # Used so downstream provider = cube can still resolve MP-level BEx queries
    idx_c2mp = {}
    if dfs.get("mp_df") is not None:
        mpdf = dfs["mp_df"]
        if "MULTIPROVIDER" in mpdf.columns and "PARTPROVIDER" in mpdf.columns:
            for _, r in mpdf.iterrows():
                mp   = str(r["MULTIPROVIDER"]).upper().strip()
                part = str(r["PARTPROVIDER"]).upper().strip()
                if part.lower() not in ("nan","none",""):
                    idx_c2mp.setdefault(part, []).append(mp)
    dfs["_idx_cube_to_mp"] = idx_c2mp

    # ds_lookup[DATASOURCE_UPPER] = {logsys, txtsh, type_desc, applnm}
    # Built from DataSource file — enriches DATASOURCE nodes with description + source system
    idx_ds = {}
    if dfs.get("ds_df") is not None:
        ddf = dfs["ds_df"]
        for col in ddf.columns:
            ddf[col] = ddf[col].astype(str).str.strip()
        if "DATASOURCE" in ddf.columns:
            for _, r in ddf.iterrows():
                k = str(r.get("DATASOURCE","")).upper().strip()
                if k.lower() in ("nan","none",""): continue
                idx_ds[k] = {
                    "logsys":    str(r.get("LOGSYS","")).strip(),
                    "txtsh":     str(r.get("TXTSH","")).strip(),
                    "type_desc": str(r.get("TYPE_DESCRIPTION","")).strip(),
                    "applnm":    str(r.get("APPLNM","")).strip(),
                }
    dfs["_idx_ds"] = idx_ds

    # ext_lookup[COMM_STRUCT_UPPER] = {tech_name, stat}
    # Built from Extractor file — maps extractor/InfoSource name to its technical comm structure
    idx_ext = {}
    if dfs.get("ext_df") is not None:
        edf = dfs["ext_df"]
        for col in edf.columns:
            edf[col] = edf[col].astype(str).str.strip()
        # Column name may have spaces — find it by partial match
        comm_col = next((c for c in edf.columns if "COMMUNICATION STRUCTURE" in c.upper() and "TECHN" not in c.upper()), None)
        tech_col = next((c for c in edf.columns if "COMMUNICATION" in c.upper() and "TECHN" in c.upper()), None)
        stat_col = next((c for c in edf.columns if c.upper() in ("STAT","STATUS","OBJSTAT")), None)
        if comm_col:
            for _, r in edf.iterrows():
                k = str(r.get(comm_col,"")).upper().strip()
                if k.lower() in ("nan","none",""): continue
                idx_ext[k] = {
                    "tech_name": str(r.get(tech_col,"")).strip() if tech_col else "",
                    "stat":      str(r.get(stat_col,"")).strip() if stat_col else "",
                }
    dfs["_idx_ext"] = idx_ext

    return dfs

# =====================================================
# OBJECT TYPE
# =====================================================

def get_object_type(obj, dfs):
    obj = str(obj).upper().strip()
    if obj in dfs["_mp_set"]:              return "MULTIPROVIDER"
    if obj in dfs["_cube_set"]:            return "INFOCUBE"
    if obj in dfs["_adso_set"]:            return "ADSO"
    if obj.endswith("_TR"):                return "INFOSOURCE"
    if obj in dfs.get("_idx_ds", {}):      return "DATASOURCE"
    return "DATASOURCE"


def get_ds_info(obj, dfs):
    """Return enriched datasource info dict or empty dict."""
    return dfs.get("_idx_ds", {}).get(str(obj).upper().strip(), {})


def get_ext_info(obj, dfs):
    """Return enriched extractor info dict or empty dict."""
    return dfs.get("_idx_ext", {}).get(str(obj).upper().strip(), {})

# =====================================================
# BUILD UPSTREAM LINEAGE (DataSource → up to Query)
# =====================================================

def build_lineage(provider, dfs):
    """
    Build upstream lineage tree using O(1) index lookups.
    Falls back to DataFrame scans if indexes are not present
    (e.g. when called from tests or old code paths).
    """
    visited = set()
    stats   = {"transformations": set(), "update_rules": set(), "ip_count": 0}

    # Use pre-built indexes when available (O(1) per lookup)
    idx_tgt  = dfs.get("_idx_trfn_by_target", {})
    idx_ip   = dfs.get("_idx_ip", {})
    idx_mp   = dfs.get("_idx_mp_parts", {})
    idx_upd  = dfs.get("_idx_upd_by_cube", {})

    def _proc(prov, allow_revisit=False):
        pu = prov.upper().strip()
        if pu in visited and not allow_revisit:
            # Return a shallow node (no children) so the object still appears in tree
            # This handles MultiProvider parts that also appear as transformation targets
            return {
                "provider": prov, "type": get_object_type(prov, dfs),
                "children": [], "transformations": [],
                "ip_nodes": [], "infopackages": 0,
            }
        visited.add(pu)
        base_pu = pu.replace("_TR", "").strip()

        node = {
            "provider": prov, "type": get_object_type(prov, dfs),
            "children": [], "transformations": [],
            "ip_nodes": [], "infopackages": 0,
        }

        # Transformations targeting this provider
        trfn_rows = idx_tgt.get(pu, [])
        if trfn_rows:
            trs = [r["TRANID"] for r in trfn_rows if r["TRANID"].lower() not in ("nan","none","")]
            node["transformations"] = trs
            stats["transformations"].update(trs)

        # Update rules (from RSUPDINFO / Update_Info file):
        # For INFOCUBE and ADSO nodes, look up all update rules feeding this cube.
        # Each update rule gives an ISOURCE (DataSource or InfoSource) that feeds the cube.
        # Use these to find InfoPackages when no transformation chain is present.
        upd_entries = idx_upd.get(pu, [])
        node["update_rules"] = [e for e in upd_entries if e.get("objstat","ACT") == "ACT"]
        for ue in node["update_rules"]:
            if ue.get("updid",""):
                stats["update_rules"].add(ue["updid"])

        # InfoPackages feeding this provider — initialise dedup sets here so update rules
        # block and the DataSource block below both share the same dedup tracking
        raw_ips = idx_ip.get(base_pu, []) or idx_ip.get(base_pu + "_TR", [])
        seen_ip = set(); dedup_ips = []
        for ip in raw_ips:
            if ip not in seen_ip:
                seen_ip.add(ip); dedup_ips.append(ip)

        for ue in node["update_rules"]:
            isrc = ue.get("isource","").strip()
            if not isrc or isrc.lower() in ("nan","none",""):
                continue
            isrc_up = isrc.upper().replace("_TR","").strip()
            # Look up InfoPackages for this ISOURCE
            ue_ips = idx_ip.get(isrc_up, []) or idx_ip.get(isrc_up + "_TR", [])
            seen_u = set(); ue_ip_list = []
            for ip in ue_ips:
                if ip not in seen_u:
                    seen_u.add(ip); ue_ip_list.append(ip)
            ue["infopackages"] = ue_ip_list
            # Enrich the ISOURCE with extractor/datasource info
            ue["ds_info"]  = get_ds_info(isrc_up, dfs)
            ue["ext_info"] = get_ext_info(isrc_up, dfs)
            stats["ip_count"] += len(ue_ip_list)
            # ── Recurse further if this ISOURCE is itself a provider with its own
            # upstream chain (e.g. an ADSO/InfoCube fed by further Update Rules or
            # Transformations) — without this, multi-hop chains stopped after one
            # hop and later hops were silently dropped.
            # SAP convention: an ADSO/InfoCube's own export InfoSource is often
            # named with an "8" prefix (e.g. ADSO "ZPCA_001" exposes as
            # "8ZPCA_001") — resolve that back to the real object key too.
            recurse_key = isrc_up
            if recurse_key not in idx_upd and recurse_key not in idx_tgt:
                stripped = recurse_key[1:] if recurse_key.startswith("8") else None
                if stripped and (stripped in dfs.get("_adso_set", set())
                                 or stripped in dfs.get("_cube_set", set())):
                    recurse_key = stripped
            if (idx_upd.get(recurse_key) or idx_tgt.get(recurse_key)) and recurse_key != pu:
                ue["child_lineage"] = _proc(recurse_key)
            # Also add to node ip_nodes (deduplicated across all sources)
            for ip in ue_ip_list:
                if ip not in seen_ip:
                    seen_ip.add(ip); dedup_ips.append(ip)

        node["infopackages"] = len(dedup_ips)
        node["ip_nodes"]     = dedup_ips
        stats["ip_count"]   += len(dedup_ips)

        # Enrich DATASOURCE nodes with description + source system from ds_df
        if node["type"] == "DATASOURCE":
            ds_info = get_ds_info(prov, dfs)
            if ds_info:
                node["logsys"]    = ds_info.get("logsys","")
                node["txtsh"]     = ds_info.get("txtsh","")
                node["type_desc"] = ds_info.get("type_desc","")
                node["applnm"]    = ds_info.get("applnm","")

        # Enrich INFOSOURCE / extractor nodes with tech name from ext_df
        if node["type"] == "INFOSOURCE":
            ext_info = get_ext_info(base_pu, dfs)
            if ext_info:
                node["tech_name"] = ext_info.get("tech_name","")
                node["ext_stat"]  = ext_info.get("stat","")

        # MultiProvider parts
        if node["type"] == "MULTIPROVIDER":
            for part in idx_mp.get(pu, []):
                ch = _proc(part)
                if ch: node["children"].append(ch)

        # Walk transformation children
        for r in trfn_rows:
            tranid = r["TRANID"]; source = r["SOURCENAME"]
            if not tranid or tranid.lower() in ("nan","none",""): continue
            base_s = source.upper().replace("_TR","").strip()
            stats["transformations"].add(tranid)

            # Try base name AND with _TR suffix — OLTPSOURCE varies by system
            src_ips = idx_ip.get(base_s, []) or idx_ip.get(base_s + "_TR", [])
            seen_s  = set(); ip_list = []
            for ip in src_ips:
                if ip not in seen_s:
                    seen_s.add(ip); ip_list.append(ip)
            stats["ip_count"] += len(ip_list)

            node["children"].append({
                "transformation": tranid,
                "source":         source,
                "lineage":        _proc(source),
                "ip_list":        ip_list,
                "infopackages":   len(ip_list),
            })

        return node

    return _proc(str(provider).strip()), stats

# =====================================================
# BUILD DOWNSTREAM LINEAGE
# Given a DataSource / InfoSource / object → find all
# Transformations, InfoProviders and Queries that consume it
# =====================================================

def build_downstream(obj_name, dfs):
    """
    Trace downstream using O(1) index lookups.
    Chain: DataSource → Transformation_1 → InfoSource_TR
                      → Transformation_2 → InfoProvider → BEx Queries
    """
    idx_src  = dfs.get("_idx_trfn_by_source", {})
    idx_q    = dfs.get("_idx_queries", {})
    idx_c2mp = dfs.get("_idx_cube_to_mp", {})
    ip_df    = dfs.get("ip_df")

    def _resolve_queries(provider):
        """Look up queries directly on provider, then fall back to parent MultiProviders."""
        prov_up = provider.upper().strip()
        qs = idx_q.get(prov_up, [])
        if not qs:
            for mp in idx_c2mp.get(prov_up, []):
                qs = qs + idx_q.get(mp, [])
        # deduplicate preserving order
        seen_q = set(); result = []
        for q in qs:
            if q not in seen_q:
                seen_q.add(q); result.append(q)
        return result
    results  = []

    base_name = obj_name.strip()
    if base_name.upper().endswith("_TR"):
        base_name = base_name[:-3].strip()
    search_up = base_name.upper()

    seen = set()
    ips  = _infopackages_for_datasource(base_name, ip_df=ip_df, dfs=dfs)

    # Resolve exact source name in index — try multiple variants
    # Some systems store DataSource with _TR suffix, client suffix, or different case
    def _resolve_src_key(name_up):
        """Try exact, then without _TR, then with _TR, then substring match."""
        candidates = [
            name_up,
            name_up.replace("_TR",""),
            name_up + "_TR",
        ]
        for c in candidates:
            if c in idx_src:
                return c
        # Last resort: substring match in index keys
        for key in idx_src:
            if name_up in key or key in name_up:
                return key
        return None

    resolved_key = _resolve_src_key(search_up)
    t1_source_rows = idx_src.get(resolved_key, []) if resolved_key else []

    # Step 1 — Transformation_1: DataSource → target
    for r1 in t1_source_rows:
        tranid1 = r1["TRANID"]
        target1 = r1["TARGETNAME"]
        if not tranid1 or tranid1.lower() in ("nan","none",""): continue

        if target1.upper().endswith("_TR"):
            infosource = target1

            # Step 2 — Transformation_2: InfoSource_TR → InfoProvider
            t2_list = idx_src.get(target1.upper(), [])

            if not t2_list:
                key = (tranid1, "", "")
                if key not in seen:
                    seen.add(key)
                    results.append({
                        "datasource":        base_name,
                        "transformation_up": tranid1,
                        "infosource":        infosource,
                        "transformation_dn": "",
                        "provider":          "",
                        "queries":           [],
                        "infopackages":      ips,
                    })
            else:
                    for r2 in t2_list:
                        tranid2  = r2["TRANID"]
                        provider = r2["TARGETNAME"]
                        if not tranid2 or tranid2.lower() in ("nan","none",""): continue

                        queries = _resolve_queries(provider)

                        key = (tranid1, tranid2, provider.upper())
                        if key not in seen:
                            seen.add(key)
                            results.append({
                                "datasource":        base_name,
                                "transformation_up": tranid1,
                                "infosource":        infosource,
                                "transformation_dn": tranid2,
                                "provider":          provider,
                                "queries":           queries,
                                "infopackages":      ips,
                            })
        else:
            provider = target1
            queries  = _resolve_queries(provider)
            key = (tranid1, "", provider.upper())
            if key not in seen:
                seen.add(key)
                results.append({
                    "datasource":        base_name,
                    "transformation_up": tranid1,
                    "infosource":        "",
                    "transformation_dn": "",
                    "provider":          provider,
                    "queries":           queries,
                    "infopackages":      ips,
                })

    return results


def _queries_for_provider(provider, query_df=None, dfs=None):
    """O(1) lookup via index when dfs provided, else DataFrame fallback."""
    if not provider: return []
    prov_up = provider.upper().strip()
    if dfs and "_idx_queries" in dfs:
        qs = dfs["_idx_queries"].get(prov_up, [])
        if not qs and "_idx_cube_to_mp" in dfs:
            for mp in dfs["_idx_cube_to_mp"].get(prov_up, []):
                qs = qs + dfs["_idx_queries"].get(mp, [])
        seen_q = set(); result = []
        for q in qs:
            if q not in seen_q:
                seen_q.add(q); result.append(q)
        return result
    if query_df is None or "INFOPROVIDER" not in query_df.columns:
        return []
    qids = query_df[
        query_df["INFOPROVIDER"].str.upper().eq(prov_up)
    ]["QUERYID"].dropna().astype(str).tolist()
    return [q for q in qids if q.lower() != "nan"]


def _infopackages_for_datasource(datasource, ip_df=None, dfs=None):
    """O(1) lookup via index when dfs provided, else DataFrame fallback."""
    if not datasource: return []
    base = datasource.replace("_TR","").upper().strip()

    # ── Fast path: use pre-built index
    if dfs and "_idx_ip" in dfs:
        # Try exact, then with _TR, then without _TR
        raw = (dfs["_idx_ip"].get(base, []) or
               dfs["_idx_ip"].get(base + "_TR", []))
        seen = set(); result = []
        for ip in raw:
            if ip not in seen:
                seen.add(ip); result.append(ip)
        return result

    # ── Fallback: scan DataFrame directly
    if ip_df is None: return []
    ip_cols = ip_df.columns.tolist()
    # Support both OLTPSOURCE and SOURCE column names
    src_col  = "OLTPSOURCE" if "OLTPSOURCE" in ip_cols else ("SOURCE" if "SOURCE" in ip_cols else None)
    # Support both LOGDPID and LOGDPID_N
    ipid_col = "LOGDPID_N"  if "LOGDPID_N"  in ip_cols else ("LOGDPID" if "LOGDPID" in ip_cols else None)
    if not src_col or not ipid_col: return []

    rows = ip_df[ip_df[src_col].str.upper().str.strip().eq(base)]
    ips  = rows[ipid_col].dropna().astype(str).tolist()
    seen = set(); result = []
    for ip in ips:
        if ip.lower() not in ("nan","none","") and ip not in seen:
            seen.add(ip); result.append(ip)
    return result


# =====================================================
# DETECT INPUT TYPE
# =====================================================

def detect_input_type(name, dfs):
    """
    Returns 'QUERY', 'DATASOURCE', or 'UNKNOWN'.
    Uses index dictionaries for O(1) exact match,
    with substring fallback for partial names.
    """
    name_up = name.upper().strip()

    # ── 1. Exact query ID match via index
    idx_q = dfs.get("_idx_queries", {})
    # idx_queries is keyed by provider, not query ID — check query_df directly for ID match
    query_df = dfs.get("query_df")
    if query_df is not None:
        if "QUERYID" in query_df.columns:
            if query_df["QUERYID"].str.upper().str.strip().eq(name_up).any():
                return "QUERY"
        if "QUERYNAME" in query_df.columns:
            if query_df["QUERYNAME"].str.upper().str.strip().eq(name_up).any():
                return "QUERY"

    # ── 2. Exact DataSource match via transformation source index
    idx_src = dfs.get("_idx_trfn_by_source", {})
    if name_up in idx_src:
        return "DATASOURCE"

    # ── 3. Exact InfoPackage OLTPSOURCE match via IP index
    idx_ip = dfs.get("_idx_ip", {})
    if name_up in idx_ip:
        return "DATASOURCE"

    # ── 4. Substring fallback (partial names) — query first
    # regex=False is critical: query names often contain . + ( ) which are regex special chars
    if query_df is not None:
        if "QUERYID" in query_df.columns:
            if query_df["QUERYID"].str.upper().str.strip().str.contains(name_up, na=False, regex=False).any():
                return "QUERY"
        if "QUERYNAME" in query_df.columns:
            if query_df["QUERYNAME"].str.upper().str.strip().str.contains(name_up, na=False, regex=False).any():
                return "QUERY"

    # ── 5. Substring fallback for DataSource
    for key in idx_src:
        if name_up in key:
            return "DATASOURCE"
    for key in idx_ip:
        if name_up in key:
            return "DATASOURCE"

    return "UNKNOWN"

# =====================================================
# COMPLEXITY
# =====================================================

def compute_complexity(stats):
    t = len(stats["transformations"])
    u = len(stats.get("update_rules", set()))
    i = stats["ip_count"]
    score = (t * 5) + (u * 3) + i
    cx = "LOW" if score < 10 else ("MEDIUM" if score < 30 else "HIGH")
    return t, u, i, score, cx

# =====================================================
# FLATTEN TREE → VISUAL ROWS
# =====================================================

def flatten_tree(node, rows=None, depth=0):
    if rows is None: rows = []
    if not node: return rows
    rows.append({"type": node.get("type","DATASOURCE"), "name": node["provider"], "depth": depth})
    # Show DataSource enrichment — source system + description
    if node.get("logsys"):
        rows.append({"type":"SRCSYS",  "name": "Source System : " + node["logsys"],  "depth": depth+1})
    if node.get("txtsh"):
        rows.append({"type":"SRCSYS",  "name": "Description   : " + node["txtsh"],   "depth": depth+1})
    if node.get("type_desc"):
        rows.append({"type":"SRCSYS",  "name": "Type          : " + node["type_desc"],"depth": depth+1})
    for tr in node.get("transformations",[]): rows.append({"type":"TRANSFORMATION","name":tr,"depth":depth+1})
    for ue in node.get("update_rules",[]):
        rows.append({"type":"UPDATERULE","name":"Update Rule ID: "+ue.get("updid",""),"depth":depth+1})
        if ue.get("isource",""):
            rows.append({"type":"EXTRACTOR","name":"Extractor Name: "+ue.get("isource",""),"depth":depth+2})
        # Show source system if available from ds_df
        ds_i = ue.get("ds_info",{})
        if ds_i.get("logsys"):
            rows.append({"type":"SRCSYS","name":"Source System : "+ds_i["logsys"],"depth":depth+3})
        if ds_i.get("txtsh"):
            rows.append({"type":"SRCSYS","name":"Description   : "+ds_i["txtsh"],"depth":depth+3})
        # Continue walking if this ISOURCE is itself a further provider with its
        # own upstream chain (e.g. ADSO fed by another Update Rule/Transformation)
        if ue.get("child_lineage"):
            flatten_tree(ue["child_lineage"], rows, depth+3)
    for ip in node.get("ip_nodes",[]): rows.append({"type":"INFOPACKAGE","name":ip,"depth":depth+1})
    for child in node.get("children",[]):
        if "transformation" in child:
            rows.append({"type":"TRANSFORMATION","name":child["transformation"],"depth":depth+1})
            rows.append({"type":"SOURCE","name":child["source"],"depth":depth+2})
            for ip in child.get("ip_list",[]): rows.append({"type":"INFOPACKAGE","name":ip,"depth":depth+3})
            if child.get("lineage"): flatten_tree(child["lineage"],rows,depth+2)
        else:
            flatten_tree(child,rows,depth+1)
    return rows

# =====================================================
# VISUAL RENDERER
# =====================================================

_TYPE_META = {
    "QUERY":          ("ln-query",       "🔍", "BEx Query"),
    "INFOPROVIDER":   ("ln-provider",    "🔀", "InfoProvider"),
    "MULTIPROVIDER":  ("ln-mp",          "🔀", "MultiProvider"),
    "INFOCUBE":       ("ln-cube",        "🧊", "InfoCube"),
    "ADSO":           ("ln-adso",        "🗄️",  "ADSO"),
    "TRANSFORMATION": ("ln-trfn",        "⚙️",  "Transformation"),
    "UPDATERULE":     ("ln-dtp",          "🔄", "Update Rule"),
    "EXTRACTOR":      ("ln-datasource",   "📡", "Extractor"),
    "DTP":            ("ln-dtp",         "🔁", "DTP"),
    "DATASOURCE":     ("ln-datasource",  "📡", "DataSource"),
    "SOURCE":         ("ln-datasource",  "📡", "DataSource"),
    "INFOSOURCE":     ("ln-infosource",  "📥", "InfoSource"),
    "INFOPACKAGE":    ("ln-infopackage", "📬", "InfoPackage"),
    "SOURCE SYSTEM":  ("ln-srcsys",      "🏭", "Source System"),
    "DOWNSTREAM":     ("ln-downstream",  "⬇️",  "Downstream"),
}

def render_lineage_visual(rows):
    if not rows: return
    html = []
    for row in rows:
        otype  = str(row["type"]).upper()
        depth  = min(row["depth"], 12)
        indent = depth * 18
        cls, icon, label = _TYPE_META.get(otype, ("ln-default","📦",otype))
        connector = "└─" if depth > 0 else ""
        html.append(
            f"<div class='ln-row {cls}' style='margin-left:{indent}px'>"
            f"<span class='ln-connector'>{connector}</span>"
            f"<span>{icon}</span>"
            f"<span class='ln-type'>{label}</span>"
            f"<span style='opacity:.3'>·</span>"
            f"<span class='ln-name'>{row['name']}</span>"
            f"</div>"
        )
    st.markdown("<div class='ln-wrap'>" + "\n".join(html) + "</div>", unsafe_allow_html=True)
    seen = list(dict.fromkeys(str(r["type"]).upper() for r in rows))
    legend = []
    for otype in seen:
        cls, icon, label = _TYPE_META.get(otype, ("ln-default","📦",otype))
        legend.append(f"<span class='ln-row {cls} ln-leg-item'>{icon} {label}</span>")
    st.markdown("<div class='ln-legend'>" + "".join(legend) + "</div>", unsafe_allow_html=True)


def render_downstream_visual(ds_results, dfs=None):
    if dfs is None: dfs = {"_mp_set":set(),"_cube_set":set(),"_adso_set":set()}
    """Render downstream chains as colored cards."""
    if not ds_results:
        st.info("No downstream consumers found for this object.")
        return

    st.markdown(f"<div class='ln-section-hdr'>⬇️ Downstream — {len(ds_results)} chain(s) found</div>",
                unsafe_allow_html=True)

    for i, r in enumerate(ds_results):
        rows = []
        # ── Deduplicate InfoPackages before rendering
        seen_ips = set()
        for ip in r.get("infopackages", []):
            if ip and ip not in seen_ips:
                seen_ips.add(ip)
                rows.append({"type": "INFOPACKAGE", "name": ip, "depth": 0})
        rows.append({"type": "DATASOURCE", "name": r["datasource"], "depth": 1})
        if r["transformation_up"]:
            rows.append({"type": "TRANSFORMATION", "name": r["transformation_up"], "depth": 2})
        if r["infosource"]:
            rows.append({"type": "INFOSOURCE",     "name": r["infosource"],        "depth": 3})
        if r["transformation_dn"]:
            rows.append({"type": "TRANSFORMATION", "name": r["transformation_dn"], "depth": 4})
        if r["provider"]:
            ptype = get_object_type(r["provider"], dfs)
            rows.append({"type": ptype, "name": r["provider"], "depth": 5})
        for q in r["queries"]:
            rows.append({"type": "QUERY", "name": q, "depth": 6})

        render_lineage_visual(rows)
        if i < len(ds_results) - 1:
            st.markdown("<hr style='margin:8px 0;border-color:#e2e8f0'>", unsafe_allow_html=True)

# =====================================================
# TREE → EXCEL
# =====================================================

def _write_excel_sheet(ws, headers, col_widths, rows_data, merge_specs, fills, section_label=""):
    """Helper: write headers + data rows + apply merges to a worksheet."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin        = Side(style="thin", color="BFBFBF")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill    = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    hdr_font    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font   = Font(name="Arial", size=9)
    merge_align = Alignment(horizontal="center", vertical="center")
    left_align  = Alignment(horizontal="left",   vertical="center")

    for ci, col in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci-1]
    ws.row_dimensions[1].height = 30

    for ri, row_vals in enumerate(rows_data):
        er = ri + 2
        fill = fills[ri % len(fills)]  # caller sets fill per row
        for ci, val in enumerate(row_vals["values"], 1):
            cell = ws.cell(row=er, column=ci, value=val)
            cell.font = data_font
            cell.fill = row_vals.get("fill", fill)
            cell.alignment = left_align
            cell.border = border
        ws.row_dimensions[er].height = 16

    # Apply merges
    for (er_start, er_end, cols, fill_obj) in merge_specs:
        if er_start >= er_end:
            continue
        for ci in cols:
            cl = get_column_letter(ci)
            try:
                ws.merge_cells(f"{cl}{er_start}:{cl}{er_end}")
                cell = ws.cell(row=er_start, column=ci)
                cell.fill = fill_obj
                cell.font = data_font
                cell.alignment = merge_align
                cell.border = border
            except Exception:
                pass

    ws.freeze_panes = "A2"


def _excel_styles():
    """Return commonly used openpyxl style objects."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin      = Side(style="thin", color="BFBFBF")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill  = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=9)
    ctr_align = Alignment(horizontal="center", vertical="center")
    lft_align = Alignment(horizontal="left",   vertical="center")
    return border, hdr_fill, hdr_font, hdr_align, data_font, ctr_align, lft_align

FILLS = [
    None,  # placeholder — defined inside functions using openpyxl
]

def _make_fills():
    from openpyxl.styles import PatternFill
    return [
        PatternFill("solid", start_color="D6E4F7", end_color="D6E4F7"),
        PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC"),
        PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA"),
        PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6"),
        PatternFill("solid", start_color="F4E1F7", end_color="F4E1F7"),
        PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7"),
    ]


def _write_header(ws, cols, widths):
    """Write a styled header row to ws."""
    from openpyxl.utils import get_column_letter
    border, hdr_fill, hdr_font, hdr_align, _, _, _ = _excel_styles()
    for ci, (col, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30


def _write_data_row(ws, row_idx, values, fill):
    """Write one data row with styling."""
    from openpyxl.utils import get_column_letter
    border, _, _, _, data_font, _, lft_align = _excel_styles()
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        cell.font = data_font; cell.fill = fill
        cell.alignment = lft_align; cell.border = border
    ws.row_dimensions[row_idx].height = 16


def _apply_merge(ws, er_start, er_end, cols, fill):
    """Merge cells and re-style the top-left cell."""
    from openpyxl.utils import get_column_letter
    border, _, _, _, data_font, ctr_align, _ = _excel_styles()
    for ci in cols:
        cl = get_column_letter(ci)
        try:
            ws.merge_cells(f"{cl}{er_start}:{cl}{er_end}")
            cell = ws.cell(row=er_start, column=ci)
            cell.fill = fill; cell.font = data_font
            cell.alignment = ctr_align; cell.border = border
        except Exception:
            pass


def _hierarchical_merge(ws, row_start, row_end, col_start, col_end):
    """
    Merge vertically consecutive, equal, non-empty cells column by column
    (left → right), but only within a run that is already a single merged
    (or singleton) block in the column to its left.

    This is what makes a dynamic-width lineage table readable: identical
    values (same BEx Name, same Root Provider, same shared hop) collapse
    into one merged cell spanning their rows instead of being repeated on
    every single row — while rows from a genuinely different branch that
    happen to share the same text later on are never merged together,
    because their earlier columns didn't match.
    """
    from openpyxl.utils import get_column_letter
    _, _, _, _, _, ctr_align, _ = _excel_styles()
    parent_ranges = [(row_start, row_end)]
    for col in range(col_start, col_end + 1):
        cl = get_column_letter(col)
        new_ranges = []
        for (rs, re) in parent_ranges:
            if re <= rs:
                new_ranges.append((rs, re))
                continue
            run_start = rs
            prev_val  = ws.cell(row=rs, column=col).value
            for r in range(rs + 1, re + 1):
                val = ws.cell(row=r, column=col).value
                if val != prev_val or not val:
                    if r - 1 > run_start:
                        try:
                            ws.merge_cells(f"{cl}{run_start}:{cl}{r-1}")
                            ws.cell(row=run_start, column=col).alignment = ctr_align
                        except Exception:
                            pass
                    new_ranges.append((run_start, r - 1))
                    run_start = r
                    prev_val  = val
            if re > run_start:
                try:
                    ws.merge_cells(f"{cl}{run_start}:{cl}{re}")
                    ws.cell(row=run_start, column=col).alignment = ctr_align
                except Exception:
                    pass
            new_ranges.append((run_start, re))
        parent_ranges = new_ranges


def _build_upstream_rows(query_name, provider, tree, input_type="QUERY"):
    """
    Build flat rows for the Upstream Excel sheet.

    For QUERY input (root = InfoCube/ADSO/MultiProvider):
      Columns: Bex Name | Root Provider | Transformation | InfoSource |
               InfoSource Transformation | DataSource | InfoPackage
      Walks children: each child = one transformation chain upward.
      trfn     = Transformation (InfoProvider → InfoSource)
      source   = InfoSource (_TR)
      is_trfn  = Transformation (InfoSource → DataSource)
      ds       = DataSource
      ips      = InfoPackages for that DataSource

    For DATASOURCE input (root = DataSource):
      Simple table: DataSource | InfoPackage
      Columns 1-6 = DataSource name, col 7 = InfoPackage
    """
    FILLS = _make_fills()
    flat_rows  = []
    merge_spec = []

    if input_type == "DATASOURCE":
        # For DataSource: show InfoPackage → DataSource only
        # ip_nodes are stored on the tree node itself by build_lineage
        ips = []
        if tree:
            ips = tree.get("ip_nodes", [])
            # Also check children — build_lineage may put IPs on child nodes
            for child in tree.get("children", []):
                if "transformation" in child:
                    ips.extend(child.get("ip_list", []))
        # Deduplicate
        seen_ip = set(); ips_dedup = []
        for ip in ips:
            if ip and ip.lower() not in ("nan","") and ip not in seen_ip:
                seen_ip.add(ip); ips_dedup.append(ip)
        if not ips_dedup:
            ips_dedup = ["(no InfoPackages found)"]
        fill = FILLS[0]
        for k, ip in enumerate(ips_dedup):
            flat_rows.append({"values": [
                query_name if k == 0 else "",  # DataSource (col A)
                ip,                            # InfoPackage (col B)
            ], "fill": fill})
        if len(flat_rows) > 1:
            merge_spec.append((2, len(flat_rows)+1, [1], FILLS[0]))
        return flat_rows, merge_spec

    # ── QUERY path: recursively collect every upstream hop-chain from the
    # root provider down to its DataSource(s)/leaf, at ANY depth. This
    # replaces the old fixed 7-column walk, which silently dropped hops
    # beyond "Transformation → InfoSource → Transformation → DataSource"
    # instead of showing them — e.g. an ADSO fed by another Update Rule,
    # or additional MultiProvider parts, just didn't get a column before.
    def _collect_hop_chains(node):
        """
        Returns a list of {"hops": [...], "ips": [...]} — one entry per
        upstream branch reachable from `node`. Each hop is
        {"type": <node/edge type>, "name": <value>}. Recurses through
        Transformation children, MultiProvider parts, and Update Rules
        (including Update Rules whose ISOURCE is itself a further provider
        with its own upstream chain) to any depth.
        """
        chains = []

        for ue in node.get("update_rules", []):
            hops = [{"type": "UPDATERULE", "name": ue.get("updid", "")}]
            if ue.get("isource"):
                hops.append({"type": "EXTRACTOR", "name": ue.get("isource", "")})
            own_ips = ue.get("infopackages", [])
            child = ue.get("child_lineage")
            if child:
                sub = _collect_hop_chains(child)
                child_hop = {"type": child.get("type", "OBJECT"), "name": child.get("provider", "")}
                if sub:
                    for sc in sub:
                        chains.append({"hops": hops + [child_hop] + sc["hops"], "ips": sc["ips"] or own_ips})
                else:
                    chains.append({"hops": hops + [child_hop], "ips": child.get("ip_nodes", []) or own_ips})
            else:
                chains.append({"hops": hops, "ips": own_ips})

        for child in node.get("children", []):
            if "transformation" in child:
                hops = [{"type": "TRANSFORMATION", "name": child.get("transformation", "")}]
                lineage = child.get("lineage")
                if lineage:
                    hops.append({"type": lineage.get("type", "OBJECT"), "name": lineage.get("provider", "")})
                    sub = _collect_hop_chains(lineage)
                    if sub:
                        for sc in sub:
                            chains.append({"hops": hops + sc["hops"], "ips": sc["ips"]})
                    else:
                        ips = child.get("ip_list", []) or lineage.get("ip_nodes", [])
                        chains.append({"hops": hops, "ips": ips})
                else:
                    hops.append({"type": "SOURCE", "name": child.get("source", "")})
                    chains.append({"hops": hops, "ips": child.get("ip_list", [])})
            else:
                hops = [{"type": child.get("type", "OBJECT"), "name": child.get("provider", "")}]
                sub = _collect_hop_chains(child)
                if sub:
                    for sc in sub:
                        chains.append({"hops": hops + sc["hops"], "ips": sc["ips"]})
                else:
                    chains.append({"hops": hops, "ips": child.get("ip_nodes", [])})

        return chains

    chains = _collect_hop_chains(tree) if tree else []
    if not chains:
        return [], 0

    max_hops = max(len(c["hops"]) for c in chains)

    for gi, c in enumerate(chains):
        fill = FILLS[gi % len(FILLS)]
        ips  = c["ips"] if c["ips"] else [""]
        for ip in ips:
            row_vals = [query_name, provider]
            for hi in range(max_hops):
                if hi < len(c["hops"]):
                    h = c["hops"][hi]
                    label = _TYPE_META.get(str(h["type"]).upper(), (None, None, str(h["type"])))[2]
                    row_vals.append(f"{label}: {h['name']}" if h.get("name") else "")
                else:
                    row_vals.append("")
            row_vals.append(ip)
            flat_rows.append({"values": row_vals, "fill": fill})

    return flat_rows, max_hops


def _write_upstream_sheet(ws, query_name, provider, tree, input_type="QUERY"):
    if input_type == "DATASOURCE":
        UP_COLS   = ["DataSource", "InfoPackage (feeds this DataSource)"]
        UP_WIDTHS = [28, 42]
        _write_header(ws, UP_COLS, UP_WIDTHS)
        flat_rows, merge_specs = _build_upstream_rows(query_name, provider, tree, input_type)
        for ri, row in enumerate(flat_rows):
            _write_data_row(ws, ri+2, row["values"], row["fill"])
        for (er_s, er_e, cols, fill) in merge_specs:
            _apply_merge(ws, er_s, er_e, cols, fill)
        ws.freeze_panes = "A2"
        return

    # ── QUERY path: column count grows to fit however deep the chains go —
    # no more silently truncating extra hops or extra MultiProvider parts.
    flat_rows, max_hops = _build_upstream_rows(query_name, provider, tree, input_type)

    if not flat_rows:
        ws.cell(row=1, column=1, value="No lineage data found")
        return

    col_names  = ["BEx Name", "Root Provider"] + [f"Step {i+1}" for i in range(max_hops)] + ["InfoPackage"]
    col_widths = [28, 20] + [34]*max_hops + [38]

    _write_header(ws, col_names, col_widths)

    for ri, row in enumerate(flat_rows):
        _write_data_row(ws, ri+2, row["values"], row["fill"])

    # Merge repeated values vertically instead of duplicating them row after
    # row (BEx Name, Root Provider, and any shared hop), respecting hierarchy
    # so unrelated branches that happen to share text later on never merge.
    if len(flat_rows) > 1:
        _hierarchical_merge(ws, 2, len(flat_rows)+1, 1, len(col_names))

    ws.freeze_panes = "A2"


def _write_downstream_sheet(ws, downstream):
    """Write downstream chains — one sheet for all chains."""
    FILLS = _make_fills()
    # Downstream columns — detect which have data
    _DS_ALL = [
        ("InfoPackage",                            0, 35),
        ("DataSource",                             1, 22),
        ("Transformation 1 (DS→InfoSource)",       2, 40),
        ("InfoSource",                             3, 22),
        ("Transformation 2 (InfoSource→Provider)", 4, 40),
        ("InfoProvider",                           5, 20),
        ("BEx Query Name",                         6, 38),
    ]
    # Build flat rows first
    flat_dn  = []
    merge_dn = []
    grp_idx  = -1

    for chain in downstream:
        qs  = chain["queries"]              if chain["queries"]              else ["(no queries found)"]
        ips = chain.get("infopackages", []) if chain.get("infopackages")    else [""]
        grp_idx = (grp_idx + 1) % len(FILLS)
        fill    = FILLS[grp_idx]
        c_start = len(flat_dn)
        total   = max(len(ips), len(qs))
        for ri in range(total):
            flat_dn.append({"values": [
                ips[ri]                    if ri < len(ips) else "",   # 0 InfoPackage
                chain["datasource"]        if ri == 0 else "",         # 1 DataSource
                chain["transformation_up"] if ri == 0 else "",         # 2 Trfn1
                chain["infosource"]        if ri == 0 else "",         # 3 InfoSource
                chain["transformation_dn"] if ri == 0 else "",         # 4 Trfn2
                chain["provider"]          if ri == 0 else "",         # 5 InfoProvider
                qs[ri]                     if ri < len(qs)  else "",   # 6 BEx Query
            ], "fill": fill})
        c_end = len(flat_dn) - 1
        if c_end > c_start:
            er_s = c_start + 2; er_e = c_end + 2
            merge_dn.append((er_s, er_e, [2,3,4,5,6], fill))

    # Detect active columns
    active_idx = set()
    for row in flat_dn:
        for ci, val in enumerate(row["values"]):
            if val and str(val).strip():
                active_idx.add(ci)
    # Always include DataSource (1) and BEx Query (6)
    active_idx.update([1, 6])

    dyn_cols   = [(name, orig_ci, w) for name, orig_ci, w in _DS_ALL if orig_ci in active_idx]
    col_names  = [c[0] for c in dyn_cols]
    col_widths = [c[2] for c in dyn_cols]
    col_map    = {orig_ci: new_ci for new_ci, (_, orig_ci, _) in enumerate(dyn_cols)}

    _write_header(ws, col_names, col_widths)

    for ri, row in enumerate(flat_dn):
        orig_vals = row["values"]
        new_vals  = [orig_vals[orig_ci] if orig_ci < len(orig_vals) else ""
                     for (_, orig_ci, _) in dyn_cols]
        _write_data_row(ws, ri+2, new_vals, row["fill"])

    for (er_s, er_e, orig_cols, fill) in merge_dn:
        new_cols = [col_map[c-1]+1 for c in orig_cols if (c-1) in col_map]
        if new_cols:
            _apply_merge(ws, er_s, er_e, new_cols, fill)

    ws.freeze_panes = "A2"


def _write_all_bex_sheet(ws, downstream):
    """One row per unique (provider, query) — all BEx names related to source."""
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
    BEX_COLS   = ["DataSource","InfoProvider","BEx Query Name"]
    BEX_WIDTHS = [25, 22, 45]
    _write_header(ws, BEX_COLS, BEX_WIDTHS)
    seen = set(); bex_rows = []
    for chain in downstream:
        for q in chain["queries"]:
            key = (chain["provider"].upper(), q.upper())
            if key not in seen:
                seen.add(key)
                bex_rows.append((chain["datasource"], chain["provider"], q))
    bex_rows.sort(key=lambda x: (x[1], x[2]))
    alt = [PatternFill("solid", start_color="EBF5FF", end_color="EBF5FF"),
           PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")]
    for ri, (ds, prov, q) in enumerate(bex_rows):
        _write_data_row(ws, ri+2, [ds, prov, q], alt[ri % 2])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(3)}{max(len(bex_rows)+1,2)}"


def _write_summary_sheet(ws, entries):
    """
    entries = list of (label, value) tuples.
    """
    from openpyxl.styles import Font, Border, Side
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold   = Font(name="Arial", bold=True, size=10)
    norm   = Font(name="Arial", size=10)
    for ri, (label, val) in enumerate(entries, 1):
        c1 = ws.cell(row=ri, column=1, value=label)
        c2 = ws.cell(row=ri, column=2, value=val)
        c1.font = bold; c2.font = norm
        c1.border = border; c2.border = border
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50


# ── Main Excel builder ──────────────────────────────────────────────────────

def tree_to_excel(query_name, provider, tree, downstream=None, input_type="QUERY"):
    """
    Single entry Excel export — one sheet only, no extras.

    BEx Query  (input_type=QUERY)      → one sheet "Upstream Lineage"
    DataSource (input_type=DATASOURCE) → one sheet "Downstream"
    """
    import openpyxl
    wb  = openpyxl.Workbook()
    ws  = wb.active

    if input_type == "DATASOURCE":
        ws.title = "Downstream"
        chains = downstream if downstream else []
        _write_downstream_sheet(ws, chains)
    else:
        ws.title = "Upstream Lineage"
        _write_upstream_sheet(ws, query_name, provider, tree, input_type)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def batch_to_excel(results):
    """
    Batch Excel — one sheet per entry, no combined/summary sheets.

    BEx Query  results → sheet named after query  (upstream format)
    DataSource results → sheet named after source (downstream format)
    """
    import openpyxl
    wb    = openpyxl.Workbook()
    first = True

    for r in results:
        # Sheet name: max 31 chars, no special chars
        raw_name   = r["query"]
        safe_name  = "".join(c if c.isalnum() or c in "_-" else "_" for c in raw_name)[:28]
        ws = wb.active if first else wb.create_sheet(safe_name)
        if first:
            ws.title = safe_name
            first    = False

        if r.get("input_type") == "DATASOURCE":
            chains = r.get("downstream") or []
            _write_downstream_sheet(ws, chains)
        else:
            _write_upstream_sheet(ws, r["query"], r["provider"],
                                  r["tree"], r.get("input_type","QUERY"))

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# =====================================================
# TREE → TEXT
# =====================================================

def tree_to_text(tree):
    """Internal use — compact text for AI context (150 line limit)."""
    lines = []
    B, L, P = "├── ", "└── ", "│   "
    def _walk(node, level=0):
        if not node: return
        pad = "    " * level
        lines.append(f"{pad}[{node['type']}]  {node['provider']}")
        if node.get("transformations"):
            lines.append(f"{pad}{B}Transformations : {len(node['transformations'])}")
            for i,tr in enumerate(node["transformations"]):
                lines.append(f"{pad}{P}  {L if i==len(node['transformations'])-1 else B}{tr}")
        if node.get("update_rules"):
            lines.append(f"{pad}{B}Update Rules : {len(node['update_rules'])}")
            for i,ue in enumerate(node["update_rules"]):
                c = L if i==len(node["update_rules"])-1 else B
                lines.append(f"{pad}{P}  {c}Update Rule ID : {ue.get('updid','')}")
                lines.append(f"{pad}{P}   {SP}Extractor Name : {ue.get('isource','')}")
        ip_nodes = node.get("ip_nodes",[])
        if ip_nodes:
            lines.append(f"{pad}{B}InfoPackages : {len(ip_nodes)}")
            for i,ip in enumerate(ip_nodes):
                lines.append(f"{pad}{P}  {L if i==len(ip_nodes)-1 else B}{ip}")
        for child in node.get("children",[]):
            if "transformation" in child:
                lines.append(f"{pad}{B}Transformation : {child['transformation']}")
                lines.append(f"{pad}{P}  {B}Source       : {child['source']}")
                il = child.get("ip_list",[])
                lines.append(f"{pad}{P}  {L}InfoPackages : {len(il)}")
                for i,ip in enumerate(il):
                    lines.append(f"{pad}{P}      {L if i==len(il)-1 else B}{ip}")
                if child.get("lineage"): _walk(child["lineage"],level+1)
            else:
                _walk(child,level+1)
    _walk(tree)
    return "\n".join(lines)


def tree_to_txt_download(query_name, provider, tree, stats, downstream=None, input_type="QUERY"):
    """
    Generate a full formatted TXT output matching the exact tree structure
    requested by the user — with ├──, └──, │ connectors at every level.
    """
    from datetime import datetime
    B  = "├── "
    L  = "└── "
    P  = "│   "
    SP = "    "

    lines = []

    # ── Header
    lines += [
        "=" * 70,
        f"  BW LINEAGE REPORT",
        f"  Generated : {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"  Input     : {query_name}",
        f"  Type      : {input_type}",
        "=" * 70,
        "",
    ]

    if input_type == "QUERY":
        # ── Upstream tree
        def _write_node(node, pad="", is_last=True):
            if not node:
                return
            connector = L if is_last else B
            ptype     = node.get("type", "UNKNOWN")
            pname     = node.get("provider", "")
            lines.append(f"{pad}{connector if pad else ''}{pname} ({ptype})")

            child_pad = pad + (SP if is_last else P)

            # Transformations list
            trfns = node.get("transformations", [])
            if trfns:
                lines.append(f"{child_pad}{B}Transformations : {len(trfns)}")
                for tr in trfns:
                    lines.append(f"{child_pad}{P}  {L}{tr}")

            # Update rules
            ues = node.get("update_rules", [])
            if ues:
                lines.append(f"{child_pad}{B}Update Rules : {len(ues)}")
                for ue in ues:
                    lines.append(f"{child_pad}{P}  {B}Update Rule ID : {ue.get('updid','')}")
                    lines.append(f"{child_pad}{P}   {SP}Extractor Name : {ue.get('isource','')}")

            # InfoPackages on this node
            ips = node.get("ip_nodes", [])
            if ips:
                lines.append(f"{child_pad}{B}InfoPackages : {len(ips)}")
                for ip in ips:
                    lines.append(f"{child_pad}{P}  {L}{ip}")

            # Children (transformation chains)
            children = node.get("children", [])
            for ci, child in enumerate(children):
                is_last_child = (ci == len(children) - 1)
                if "transformation" in child:
                    tc = L if is_last_child else B
                    tranid = child.get("transformation", "")
                    source = child.get("source", "")
                    il     = child.get("ip_list",  [])
                    cp     = child_pad + (SP if is_last_child else P)

                    lines.append(f"{child_pad}{tc}Transformation : {tranid}")
                    lines.append(f"{cp}  {B}Source       : {source}")
                    lines.append(f"{cp}  {L}InfoPackages : {len(il)}")
                    for ip in il:
                        lines.append(f"{cp}      {L}{ip}")
                    if child.get("lineage"):
                        _write_node(child["lineage"], pad=cp + "  ", is_last=True)
                else:
                    _write_node(child, pad=child_pad, is_last=is_last_child)

        _write_node(tree, pad="", is_last=True)

    elif input_type == "DATASOURCE" and downstream:
        # ── Downstream chains
        lines.append(f"DOWNSTREAM — {len(downstream)} chain(s)")
        lines.append("-" * 70)
        for ci, chain in enumerate(downstream):
            lines.append("")
            lines.append(f"Chain {ci+1}:")
            ips = chain.get("infopackages", [])
            for ip in ips:
                lines.append(f"  {B}InfoPackage         : {ip}")
            lines.append(f"  {B}DataSource          : {chain['datasource']}")
            if chain.get("transformation_up"):
                lines.append(f"  {B}Transformation 1    : {chain['transformation_up']}")
            if chain.get("infosource"):
                lines.append(f"  {B}InfoSource          : {chain['infosource']}")
            if chain.get("transformation_dn"):
                lines.append(f"  {B}Transformation 2    : {chain['transformation_dn']}")
            if chain.get("provider"):
                lines.append(f"  {B}InfoProvider        : {chain['provider']}")
            qs = chain.get("queries", [])
            if qs:
                lines.append(f"  {L}BEx Queries ({len(qs)}):")
                for q in qs:
                    lines.append(f"      {L}{q}")

    # ── Summary
    lines += [
        "",
        "=" * 70,
        "  SUMMARY",
        "=" * 70,
    ]
    t = len(stats.get("transformations", set()))
    u = len(stats.get("update_rules", set()))
    i = stats.get("ip_count", 0)
    score = (t * 5) + (u * 3) + i
    cx = "LOW" if score < 10 else ("MEDIUM" if score < 30 else "HIGH")
    lines += [
        f"  Transformations : {t}",
        f"  Update Rules    : {u}",
        f"  InfoPackages    : {i}",
        f"  Score           : {score}  ({cx})",
        "",
    ]

    return "\n".join(lines).encode("utf-8")

# =====================================================
# API KEY + AI CALLERS
# =====================================================

def get_api_key():
    try:
        return st.secrets.get("ANTHROPIC_API_KEY","")
    except:
        return ""

import requests, json, re as _re

def call_claude(system, user, max_tokens=900):
    key = get_api_key()
    if not key: return "[ERROR: No API key — paste it in the sidebar]"
    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"Content-Type":"application/json","x-api-key":key,"anthropic-version":"2023-06-01"},
            json={"model":"claude-sonnet-4-6","max_tokens":max_tokens,"system":system,
                  "messages":[{"role":"user","content":user}]},
            timeout=60
        )
        if r.status_code != 200: return f"[ERROR {r.status_code}: {r.text[:400]}]"
        return "".join(b.get("text","") for b in r.json().get("content",[]))
    except Exception as e: return f"[ERROR: {e}]"

def call_claude_chat(system, messages, max_tokens=600):
    key = get_api_key()
    if not key: return "[ERROR: No API key — paste it in the sidebar]"
    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"Content-Type":"application/json","x-api-key":key,"anthropic-version":"2023-06-01"},
            json={"model":"claude-sonnet-4-6","max_tokens":max_tokens,"system":system,"messages":messages},
            timeout=60
        )
        if r.status_code != 200: return f"[ERROR {r.status_code}: {r.text[:400]}]"
        return "".join(b.get("text","") for b in r.json().get("content",[]))
    except Exception as e: return f"[ERROR: {e}]"

def make_lineage_context(query_input, provider, tree, stats):
    t,u,i,score,cx = compute_complexity(stats)
    lines = tree_to_text(tree).splitlines()[:150]
    return (f"BEx Query: {query_input}\nProvider: {provider}\n"
            f"Complexity: {cx} (Score:{score})\n"
            f"Transformations:{t} | UpdateRules:{u} | InfoPackages:{i}\n\nLineage Tree:\n" + "\n".join(lines))

# =====================================================
# AI TRANSFORMATION CHAIN ANALYSIS
# Analyses all transformations found in a lineage tree
# =====================================================

def do_chain_analysis(tranids, dfs):
    """
    Given a set of TRANID strings from lineage stats,
    look up each transformation in trfn_df and ask Claude
    to analyse the full chain together.
    Returns a dict: {tranid: {sourcename, targetname, ai_summary}}
    plus an overall chain narrative.
    """
    trfn_df = dfs.get("trfn_df")
    if trfn_df is None or not tranids:
        return {}

    # Build context for each transformation
    chain_ctx = []
    results   = {}
    for tid in sorted(tranids):
        rows = trfn_df[trfn_df["TRANID"].str.upper().str.strip().eq(tid.upper().strip())]
        if rows.empty:
            continue
        row = rows.iloc[0]
        src = str(row.get("SOURCENAME","")).strip()
        tgt = str(row.get("TARGETNAME","")).strip()
        stype = str(row.get("SOURCETYPE","")).strip()
        ttype = str(row.get("TARGETTYPE","")).strip()
        results[tid] = {"sourcename": src, "targetname": tgt,
                        "sourcetype": stype, "targettype": ttype, "ai": ""}
        chain_ctx.append(
            f"  TRANID: {tid}\n"
            f"  Source: {src} (type: {stype})\n"
            f"  Target: {tgt} (type: {ttype})"
        )

    if not chain_ctx:
        return {}

    # Overall chain narrative
    ctx_str = "\n\n".join(chain_ctx)
    n = len(chain_ctx)
    overall = call_claude(
        "You are an SAP BW/4HANA migration expert.",
        (
            f"Analyse these {n} transformations that form a lineage chain.\n\n"
            f"{ctx_str}\n\n"
            "Provide:\n"
            "1. A plain-English description of the full data flow (2-3 sentences)\n"
            "2. Key migration risks for this transformation chain\n"
            "3. Recommended migration order (which to migrate first)\n"
            "4. Estimated complexity (LOW/MEDIUM/HIGH) with brief reason\n"
            "Keep the total response under 200 words."
        ),
        max_tokens=500
    )
    results["_overall"] = overall
    return results


def do_single_trfn_ai(tranid, sourcename, targetname, sourcetype, targettype):
    """Quick AI summary for a single transformation."""
    return call_claude(
        "You are an SAP BW/4HANA migration expert.",
        (
            f"Briefly explain this SAP BW transformation (2-3 sentences max):\n"
            f"TRANID: {tranid}\n"
            f"Source: {sourcename} (type: {sourcetype})\n"
            f"Target: {targetname} (type: {targettype})\n"
            "What does it likely do and what is the migration consideration?"
        ),
        max_tokens=200
    )

# =====================================================
# NOT-FOUND REPORT EXCEL
# =====================================================

def build_not_found_excel(errors, queries_to_run, dfs):
    """
    Build an Excel report of queries/objects not found in metadata.
    Includes: input value, reason, suggestions.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Not Found"

    thin      = Side(style="thin", color="BFBFBF")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill  = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=9)
    lft_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    err_fill  = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
    ok_fill   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")

    COLS   = ["#", "Input Value", "Reason Not Found", "Closest Match in Metadata", "Suggestion"]
    WIDTHS = [5,   40,            40,                  45,                           55]

    for ci, (col, w) in enumerate(zip(COLS, WIDTHS), 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 28

    query_df = dfs.get("query_df")
    idx_src  = dfs.get("_idx_trfn_by_source", {})
    idx_ip   = dfs.get("_idx_ip", {})

    for ri, err in enumerate(errors):
        excel_row = ri + 2
        val      = err["value"]
        val_up   = val.upper().strip()
        reason   = err["reason"]
        closest  = ""
        suggest  = ""

        # Try to find closest match and suggest what to do
        if reason == "QUERY_NOT_IN_METADATA":
            reason_txt = "Query ID/name not found in Query metadata file"
            # Look for partial matches in query_df
            if query_df is not None and "QUERYID" in query_df.columns:
                parts = query_df[
                    query_df["QUERYID"].str.upper().str.contains(val_up[:8], na=False)
                ]["QUERYID"].head(3).tolist()
                if parts:
                    closest = ", ".join(parts)
                    suggest = f"Did you mean one of these? Check spelling or use partial name search."
                else:
                    suggest = "Verify the Query ID exists in your Query metadata file (QUERYID column)."

        elif reason == "NO_QUERY_METADATA":
            reason_txt = "Query metadata file not uploaded"
            suggest    = "Upload the Query file (QUERYID · QUERYNAME · INFOPROVIDER) in the upload gate."

        elif reason == "UNKNOWN_TYPE":
            reason_txt = "Not found as Query, DataSource, or InfoPackage in any metadata file"
            # Check for partial matches in transformation sources
            partial_src = [k for k in idx_src if val_up[:6] in k][:3]
            partial_ip  = [k for k in idx_ip  if val_up[:6] in k][:3]
            if partial_src:
                closest = "DataSource candidates: " + ", ".join(partial_src)
                suggest = "Try entering the full DataSource name from the Transformation file."
            elif partial_ip:
                closest = "InfoPackage candidates: " + ", ".join(partial_ip)
                suggest = "Try entering the full OLTPSOURCE name from the InfoPackage file."
            else:
                suggest = "Check spelling. Ensure the relevant metadata files are uploaded."

        else:
            reason_txt = reason
            suggest    = "Check the metadata files and verify this object exists."

        fill = err_fill
        for ci, val_cell in enumerate([ri+1, val, reason_txt, closest, suggest], 1):
            cell = ws.cell(row=excel_row, column=ci, value=val_cell)
            cell.font = data_font; cell.fill = fill
            cell.alignment = lft_align; cell.border = border
        ws.row_dimensions[excel_row].height = 30

    # ── Summary sheet
    ws2 = wb.create_sheet("Summary")
    bold = Font(name="Arial", bold=True, size=10)
    norm = Font(name="Arial", size=10)
    summary = [
        ("Total Queries Input",      len(queries_to_run)),
        ("Successfully Traced",       len(queries_to_run) - len(errors)),
        ("Not Found",                 len(errors)),
        ("Success Rate",              f"{round((len(queries_to_run)-len(errors))/max(len(queries_to_run),1)*100)}%"),
        ("",""),
        ("Reason Breakdown",""),
    ]
    reason_counts = {}
    for e in errors:
        reason_counts[e["reason"]] = reason_counts.get(e["reason"], 0) + 1
    for reason, count in reason_counts.items():
        summary.append((f"  {reason}", count))

    for ri2, (label, val) in enumerate(summary, 1):
        ws2.cell(row=ri2, column=1, value=label).font = bold
        ws2.cell(row=ri2, column=2, value=val).font   = norm
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 20

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# =====================================================
# SESSION STATE
# =====================================================

for _k, _v in [
    ("result",None),("data",None),("last_file",None),
    ("lin_tree",None),("lin_stats",None),
    ("lin_query",None),("lin_error",None),
    ("lin_provider",""),("lin_results",None),
    ("lin_downstream",None),("lin_input_type",""),
    ("lin_chain_analysis",None),
    ("queries_pending",[]),("last_qfile",None),
    ("trfn_results",[]),("trfn_processing",False),
    # Upload gate
    ("meta_uploaded",{}),("meta_ready",False),("dfs_cache",None),
    ("last_zip",None),("zip_messages",[]),
]:
    if _k not in st.session_state:
        st.session_state[_k] = _v

# =====================================================
# HEADER
# =====================================================

st.markdown("<div class='hero-title'>⚙️ BW Transformation Analyzer</div>", unsafe_allow_html=True)
st.markdown("<div class='hero-sub'>AI-powered BW Transformation Analysis & Full Lineage Explorer</div>", unsafe_allow_html=True)
st.markdown("---")

# =====================================================
# METADATA UPLOAD GATE
# Shown on first launch. Users upload all 7 files here.
# After clicking Proceed, dfs is built from the uploads
# and stored in session_state so it survives reruns.
# st.stop() ensures nothing below renders until ready.
# =====================================================

if not st.session_state.meta_ready:

    st.markdown("### 📁 Upload Metadata ZIP File")
    st.markdown(
        "Put all your SAP BW metadata CSV files into one ZIP and upload it here. "
        "The tool auto-detects each file type by its columns — no renaming needed. "
        "All data is filtered to **OBJVERS = A** (active versions only)."
    )

    # ── What to include info box
    with st.expander("📋 What files to include in the ZIP?"):
        st.markdown("Add any of these exports from SAP BW (SE16 / RSA1) to your ZIP:")
        for key, meta in META_FILES.items():
            req = "★ Required" if meta["required"] else "Optional"
            colour = "#dc2626" if meta["required"] else "#64748b"
            st.markdown(
                f"<span style='color:{colour};font-size:12px'>{req}</span> "
                f"**{meta['label']}** — columns: `{meta['desc']}`",
                unsafe_allow_html=True
            )

    st.markdown("")

    zip_file = st.file_uploader(
        "Upload ZIP file", type=["zip"],
        key="zip_upload", label_visibility="collapsed"
    )

    if zip_file:
        if st.session_state.get("last_zip") != zip_file.name:
            # New ZIP — parse and detect files
            st.session_state.last_zip = zip_file.name
            with st.spinner("📂 Extracting and detecting files…"):
                uploads, messages = parse_zip_upload(zip_file)
                st.session_state.meta_uploaded = uploads
                st.session_state["zip_messages"] = messages

        # Show results
        messages = st.session_state.get("zip_messages", [])
        for msg in messages:
            if msg.startswith("✅"):
                st.success(msg)
            elif msg.startswith("⚠️"):
                st.warning(msg)
            else:
                st.error(msg)

        uploads      = st.session_state.meta_uploaded
        req_keys     = [k for k,v in META_FILES.items() if v["required"]]
        found_keys   = set(uploads.keys())
        missing      = [META_FILES[k]["label"] for k in req_keys if k not in found_keys]
        required_done = len(missing) == 0

        st.markdown("---")
        sc1, sc2, sc3 = st.columns([3, 1, 1])
        with sc1:
            if required_done:
                st.success(f"✅ {len(found_keys)}/7 files detected — all required files found")
            else:
                st.warning(f"⚠️ Missing required: {' · '.join(missing)}")
        with sc2:
            if st.button("▶️ Proceed", type="primary",
                         disabled=not required_done, use_container_width=True):
                with st.spinner("Processing metadata…"):
                    st.session_state.dfs_cache  = prepare_dfs(uploads)
                    st.session_state.meta_ready = True
                st.rerun()
        with sc3:
            if st.button("🗑️ Clear", use_container_width=True):
                st.session_state.meta_uploaded  = {}
                st.session_state.meta_ready     = False
                st.session_state.dfs_cache      = None
                st.session_state.last_zip       = None
                st.session_state.zip_messages   = []
                st.rerun()
    else:
        st.info("👆 Upload a ZIP file containing your SAP BW metadata CSV exports")

    st.stop()

# ── Files are ready — load from cache
dfs = st.session_state.dfs_cache

# =====================================================
# SIDEBAR
# =====================================================

with st.sidebar:
    st.markdown("### 📁 Metadata Files")
    for key, meta in META_FILES.items():
        df = dfs.get(key)
        if df is not None:
            st.success(f"✅ {meta['label']} ({len(df):,} rows)")
        elif meta["required"]:
            st.error(f"❌ {meta['label']} — missing")
        else:
            st.warning(f"⚠️ {meta['label']} — not uploaded")
    st.markdown("")
    if st.button("🔄 Re-upload Metadata", use_container_width=True, key="reupload_btn"):
        # Clear everything and return to upload gate
        st.session_state.meta_ready    = False
        st.session_state.dfs_cache     = None
        st.session_state.meta_uploaded = {}
        st.session_state.last_zip      = None
        st.session_state.zip_messages  = []
        for k in ["lin_tree","lin_stats","lin_query","lin_error","lin_provider",
                  "lin_results","lin_downstream","lin_input_type",
                  "queries_pending","last_qfile","result","data","last_file","trfn_results"]:
            if k in st.session_state:
                st.session_state[k] = None if k not in ("lin_provider","lin_input_type") else ""
        st.rerun()
    st.markdown("---")
    st.markdown("---")
    loaded = sum(1 for k,v in dfs.items() if not k.startswith("_") and v is not None)
    st.markdown("### 📁 Metadata Status")
    if loaded == 7:  st.success("✅ All 7 files loaded (OBJVERS = A)")
    elif loaded > 0: st.warning(f"⚠️ {loaded}/7 files loaded")
    else:            st.error("❌ No metadata files found")

# =====================================================
# LAYOUT
# =====================================================

left_col, right_col = st.columns([1,1], gap="large")

# ─────────────────────────────────────────
# LEFT — Transformation Analyzer
# ─────────────────────────────────────────
with left_col:
    st.markdown("#### 📂 Transformation Analyzer")
    lt1, lt2 = st.tabs(["📁 Single File", "📋 Batch via CSV / Excel"])
    uploaded_file = None
    batch_files   = []

    with lt1:
        st.caption("Upload one Excel file exported from your SAP BW transformation.")
        uploaded_file = st.file_uploader("Excel file", type=["xlsx","xls"],
                                         key="single_trfn", label_visibility="collapsed")
    with lt2:
        st.caption("Upload multiple transformation Excel files at once.")
        batch_files = st.file_uploader("Transformation files (multi-select)", type=["xlsx","xls"],
                                       accept_multiple_files=True, key="batch_trfn",
                                       label_visibility="collapsed")
        if batch_files:
            st.success(f"✅ {len(batch_files)} file(s) selected")

    def _analyze_one(uf):
        try:
            uf.seek(0)
            d = process_transformation(uf)
            r = analyze_code(d["code"], d["summary"])
            return {"file":uf.name,"data":d,"result":r,"error":None}
        except Exception as e:
            return {"file":uf.name,"data":None,"result":None,"error":str(e)}

    if uploaded_file:
        if uploaded_file.name != st.session_state.last_file:
            st.session_state.last_file = uploaded_file.name
            st.session_state.result    = None
            st.session_state.data      = None
            st.session_state.trfn_results = []
        if st.session_state.result is None:
            with st.spinner("🤖 Analyzing transformation…"):
                out = _analyze_one(uploaded_file)
            if out["error"]: st.error(f"❌ {out['error']}")
            else:
                st.session_state.data   = out["data"]
                st.session_state.result = out["result"]

    if batch_files:
        if st.button("▶️ Analyze All Files", type="primary", key="batch_run"):
            st.session_state.trfn_results = []
            prog = st.progress(0, text="Analyzing batch…")
            for i, uf in enumerate(batch_files):
                prog.progress(i/len(batch_files), text=f"Analyzing: {uf.name}")
                st.session_state.trfn_results.append(_analyze_one(uf))
            prog.progress(1.0, text="Done!")

    result = st.session_state.result
    data   = st.session_state.data

    if result and not st.session_state.trfn_results:
        cx     = str(result.get("complexity","UNKNOWN")).upper()
        cx_cls = {"LOW":"cx-LOW","MEDIUM":"cx-MEDIUM","HIGH":"cx-HIGH"}.get(cx,"")
        _, mid, _ = st.columns([1,1,1])
        with mid:
            st.markdown(f"<div class='metric-card'><h3>Complexity</h3><h2 class='{cx_cls}'>{cx}</h2></div>",
                        unsafe_allow_html=True)
        st.markdown("")
        m1,m2,m3 = st.columns(3)
        m1.metric("📄 Code Lines", result.get("code_lines","—"))
        m2.metric("🗂️ Z-Tables",   len(result.get("ztables",[])) if result.get("ztables") else 0)
        m3.metric("🧠 Confidence", result.get("confidence","—"))
        st.markdown("")
        t1,t2,t3,t4 = st.tabs(["📋 Overview","🤖 AI Analysis","🗂️ Z-Tables","📊 Summary Table"])
        with t1:
            st.markdown(f"<div class='ai-box'>{result.get('overview','No overview.')}</div>", unsafe_allow_html=True)
        with t2:
            st.markdown(result.get("reason","No analysis returned."))
            if st.button("🔄 Regenerate", key="regen"):
                st.session_state.result = None; st.rerun()
        with t3:
            if result.get("ztables"):
                st.info(result.get("ztable_summary",""))
                for zt in result["ztables"]:
                    st.markdown(f"<div class='ztable-card'>📦 <b>{zt}</b></div>", unsafe_allow_html=True)
            else:
                st.success("✅ No custom Z-Tables detected.")
        with t4:
            if data and data.get("summary") is not None:
                st.dataframe(data["summary"], use_container_width=True)
            else:
                st.info("No summary available.")
        st.markdown("---")
        report_df = pd.DataFrame([{
            "File":            uploaded_file.name if uploaded_file else "",
            "Complexity":      result.get("complexity",""),
            "Overview":        result.get("overview",""),
            "AI Analysis":     result.get("reason",""),
            "Z-Tables":        ", ".join(result.get("ztables",[])) if result.get("ztables") else "",
            "Z-Table Summary": result.get("ztable_summary",""),
        }])
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            report_df.to_excel(writer, index=False, sheet_name="Analysis")
            if data and data.get("summary") is not None:
                data["summary"].to_excel(writer, index=False, sheet_name="Transformation Summary")
        buf.seek(0)
        st.download_button(
            "📥 Download Report (.xlsx)", data=buf,
            file_name=f"bw_analysis_{uploaded_file.name if uploaded_file else 'report'}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    if st.session_state.trfn_results:
        st.markdown(f"#### Batch Results — {len(st.session_state.trfn_results)} files")
        summary_rows = []
        for r in st.session_state.trfn_results:
            if r["error"]:
                summary_rows.append({"File":r["file"],"Complexity":"ERROR","Z-Tables":"","Error":r["error"]})
            else:
                res = r["result"]
                summary_rows.append({"File":r["file"],"Complexity":res.get("complexity","—"),
                                     "Z-Tables":len(res.get("ztables",[])) if res.get("ztables") else 0,"Error":""})
        st.dataframe(pd.DataFrame(summary_rows), use_container_width=True)
        for r in st.session_state.trfn_results:
            with st.expander(f"{'❌' if r['error'] else '✅'} {r['file']}"):
                if r["error"]: st.error(r["error"])
                else:
                    res = r["result"]
                    st.markdown(f"**Complexity:** `{str(res.get('complexity','?')).upper()}`")
                    st.markdown(f"<div class='ai-box'>{res.get('overview','')}</div>", unsafe_allow_html=True)
                    if res.get("ztables"):
                        st.markdown(f"**Z-Tables:** {', '.join(res['ztables'])}")
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf,"w",zipfile.ZIP_DEFLATED) as zf:
            for r in st.session_state.trfn_results:
                if r["result"]:
                    res = r["result"]
                    df  = pd.DataFrame([{"File":r["file"],"Complexity":res.get("complexity",""),
                                         "Overview":res.get("overview",""),"AI Analysis":res.get("reason",""),
                                         "Z-Tables":", ".join(res.get("ztables",[])) if res.get("ztables") else ""}])
                    xbuf = io.BytesIO()
                    with pd.ExcelWriter(xbuf, engine="openpyxl") as writer:
                        df.to_excel(writer, index=False, sheet_name="Analysis")
                        if r["data"] and r["data"].get("summary") is not None:
                            r["data"]["summary"].to_excel(writer, index=False, sheet_name="Summary")
                    safe = "".join(c if c.isalnum() or c in "-_." else "_" for c in r["file"])
                    zf.writestr(f"analysis_{safe}.xlsx", xbuf.getvalue())
        zip_buf.seek(0)
        st.download_button(f"📥 Download All {len(st.session_state.trfn_results)} Reports (.zip)",
                           data=zip_buf, file_name="bw_batch_analysis.zip",
                           mime="application/zip", key="batch_dl")

# ─────────────────────────────────────────
# RIGHT — BW Lineage Explorer
# ─────────────────────────────────────────
with right_col:
    st.markdown("#### 🔗 BW Lineage Explorer")
    st.caption(
        "Enter a **BEx Query** name for top-down lineage, "
        "or a **DataSource / Extractor** name for bottom-up + downstream lineage."
    )

    inp_tab1, inp_tab2 = st.tabs(["✏️ Single Entry", "📄 Upload Query List"])

    with inp_tab1:
        st.caption("Query name/ID → top-down lineage.  DataSource/Extractor → bidirectional lineage.")
        single_q = st.text_input(
            "Query or DataSource name", placeholder="e.g. 0IC_C03_Q0001 or 2LIS_03_BX",
            key="single_q", label_visibility="collapsed"
        )

    with inp_tab2:
        st.caption(
            "Upload CSV/Excel — uses **QUERYID** column if found, then **QUERYNAME**, "
            "else **first column (A)** automatically. Header row is skipped."
        )
        qfile = st.file_uploader("Query list file", type=["csv","xls","xlsx","txt","tsv"],
                                 key="qfile_upload", label_visibility="collapsed")
        if qfile:
            # Only re-parse if file changed
            if st.session_state.get("last_qfile") != qfile.name:
                st.session_state.last_qfile = qfile.name
                try:
                    qfile.seek(0)
                    if qfile.name.lower().endswith((".xls",".xlsx")):
                        qf_df = pd.read_excel(qfile)
                    else:
                        # Use the same robust reader as metadata files
                        # Handles unquoted commas in description columns
                        qf_df = _read_uploaded_df(qfile)

                    if qf_df is not None and len(qf_df) > 0:
                        qf_df = qf_df.astype(str)
                        if   "QUERYID"   in qf_df.columns: raw_qs = qf_df["QUERYID"].tolist();   col_used = "QUERYID"
                        elif "QUERYNAME" in qf_df.columns: raw_qs = qf_df["QUERYNAME"].tolist(); col_used = "QUERYNAME"
                        else:                               raw_qs = qf_df.iloc[:,0].tolist();    col_used = qf_df.columns[0]
                        # Strip whitespace + skip blanks/nan values
                        # Also strip BOM chars (﻿) that appear in some CSV exports
                        parsed = [
                            q.strip().lstrip("﻿")
                            for q in raw_qs
                            if q.strip().lstrip("﻿")
                            and q.strip().lower() not in ("nan","none","#n/a","null","n/a")
                        ]
                        st.session_state.queries_pending = parsed
                        st.session_state["_qfile_col"]   = col_used
                except Exception as e:
                    st.error(f"Could not read file: {e}")

            # Show status from session state
            pending = st.session_state.get("queries_pending", [])
            col_used = st.session_state.get("_qfile_col", "column A")
            if pending:
                st.success(
                    f"✅ **{len(pending)}** quer{'y' if len(pending)==1 else 'ies'} "
                    f"loaded from column **{col_used}** — ready to trace"
                )
                with st.expander(f"Preview ({min(10,len(pending))} shown)"):
                    for q in pending[:10]: st.markdown(f"- `{q}`")
                    if len(pending) > 10: st.caption(f"… and {len(pending)-10} more")
            else:
                st.info("Upload a file to load query list.")
        else:
            # File removed — clear pending list
            if st.session_state.get("last_qfile"):
                st.session_state.queries_pending = []
                st.session_state.last_qfile = None

    # ── Build queries_to_run: single entry OR file upload
    queries_to_run = []
    if single_q.strip():
        queries_to_run = [single_q.strip()]
    elif st.session_state.get("queries_pending"):
        queries_to_run = st.session_state.queries_pending

    b1, b2 = st.columns([1,1])
    with b1: trace = st.button("▶️ Trace Lineage", type="primary", use_container_width=True)
    with b2:
        if st.button("🗑️ Clear", use_container_width=True):
            for k in ("lin_tree","lin_stats","lin_query","lin_error","lin_provider",
                      "lin_results","lin_downstream","lin_input_type"):
                st.session_state[k] = None if k not in ("lin_provider","lin_input_type") else ""
            st.session_state.queries_pending   = []
            st.session_state.last_qfile        = None
            st.session_state.lin_chain_analysis = None
            st.rerun()

    # ── Run trace
    if trace and queries_to_run:
        query_df = dfs.get("query_df")
        results  = []; errors = []  # errors = list of {value, reason}
        prog     = st.progress(0, text="Tracing…")

        for idx, q in enumerate(queries_to_run):
            prog.progress(idx / len(queries_to_run), text=f"Tracing: {q}")
            q_clean = q.strip().lstrip("﻿")
            q_up    = q_clean.upper()

            # ── Whole per-item body wrapped in try/except so that an unexpected
            # error on ANY single query (bad data, missing column, etc.) is recorded
            # as an error for that item instead of crashing the Streamlit script
            # and silently aborting the rest of the batch.
            try:
                # ── Step 1: Try to match as a QUERY first (regardless of detect_input_type)
                # This avoids the detection step being wrong and blocking a valid query.
                matched_as_query = False
                if query_df is not None:
                    mask = pd.Series(False, index=query_df.index)
                    # Exact match
                    if "QUERYID" in query_df.columns:
                        mask = mask | query_df["QUERYID"].str.upper().str.strip().eq(q_up)
                    if "QUERYNAME" in query_df.columns:
                        mask = mask | query_df["QUERYNAME"].str.upper().str.strip().eq(q_up)
                    # Substring fallback
                    if not mask.any():
                        if "QUERYID" in query_df.columns:
                            mask = mask | query_df["QUERYID"].str.upper().str.strip().str.contains(q_up, na=False, regex=False)
                        if "QUERYNAME" in query_df.columns:
                            mask = mask | query_df["QUERYNAME"].str.upper().str.strip().str.contains(q_up, na=False, regex=False)
                    matches = query_df[mask]
                    if not matches.empty:
                        matched_as_query = True
                        provider = str(matches.iloc[0]["INFOPROVIDER"]).strip()
                        tree, stats = build_lineage(provider, dfs)
                        results.append({
                            "query": q_clean, "provider": provider,
                            "tree": tree, "stats": stats,
                            "input_type": "QUERY", "downstream": None,
                        })

                if matched_as_query:
                    continue

                # ── Step 2: Try to match as a DATASOURCE
                idx_src    = dfs.get("_idx_trfn_by_source", {})
                idx_ip     = dfs.get("_idx_ip", {})
                exact_name = None

                if q_up in idx_src:
                    exact_name = q_clean
                elif q_up in idx_ip:
                    exact_name = q_clean
                else:
                    # Substring search in index keys
                    for key in idx_src:
                        if q_up in key:
                            exact_name = key; break
                    if not exact_name:
                        for key in idx_ip:
                            if q_up in key:
                                exact_name = key; break

                if exact_name:
                    # Use base name (strip _TR) for consistency
                    base_exact = exact_name.replace("_TR","").strip() if exact_name.upper().endswith("_TR") else exact_name
                    tree, stats   = build_lineage(base_exact, dfs)
                    downstream    = build_downstream(base_exact, dfs)
                    results.append({
                        "query": q_clean, "provider": base_exact,
                        "tree": tree, "stats": stats,
                        "input_type": "DATASOURCE", "downstream": downstream,
                    })
                    continue

                # ── Step 3: Not found anywhere
                errors.append({"value": q_clean, "reason": "UNKNOWN_TYPE"})

            except Exception as e:
                # Record the failure and keep going — this is the fix for batch
                # processing stopping after the first item.
                errors.append({"value": q_clean, "reason": f"ERROR: {e}"})

        prog.progress(1.0, text="Done!")
        if errors:
            err_vals = [e["value"] for e in errors]
            with st.expander(
                f"⚠️ {len(errors)} item(s) not found in metadata — click to view & download",
                expanded=True
            ):
                # Summary counts
                from collections import Counter
                reasons = Counter(e["reason"] for e in errors)
                ec1, ec2, ec3 = st.columns(3)
                ec1.metric("❌ Not Found",       len(errors))
                ec2.metric("✅ Successfully Traced", len(results))
                ec3.metric("📊 Total Input",     len(queries_to_run))

                st.markdown("")

                # Reason breakdown
                reason_labels = {
                    "QUERY_NOT_IN_METADATA": "Query ID not in metadata",
                    "NO_QUERY_METADATA":     "Query metadata file missing",
                    "UNKNOWN_TYPE":          "Unknown — not a query or DataSource",
                }
                for reason, count in reasons.items():
                    label = reason_labels.get(reason, reason)
                    st.markdown(
                        f'<div style="background:#fff1f2;border-left:4px solid #e11d48;'
                        f'padding:6px 12px;border-radius:4px;font-size:13px;margin:3px 0;color:#9f1239">'
                        f'❌ <b>{count}×</b> — {label}</div>',
                        unsafe_allow_html=True
                    )

                # Preview first 20
                st.markdown("")
                st.markdown(f"**Not found values** ({min(20,len(err_vals))} of {len(err_vals)} shown):")
                preview_cols = st.columns(2)
                for pi, ev in enumerate(err_vals[:20]):
                    preview_cols[pi % 2].code(ev)
                if len(err_vals) > 20:
                    st.caption(f"… and {len(err_vals)-20} more in the downloaded report")

                # Download button
                st.markdown("")
                err_buf = build_not_found_excel(errors, queries_to_run, dfs)
                st.download_button(
                    "⬇️ Download Not-Found Report (.xlsx)",
                    data=err_buf,
                    file_name="bw_not_found_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="err_dl"
                )

        if results:
            if len(results) == 1:
                r = results[0]
                st.session_state.lin_tree       = r["tree"]
                st.session_state.lin_stats      = r["stats"]
                st.session_state.lin_query      = r["query"]
                st.session_state.lin_provider   = r["provider"]
                st.session_state.lin_input_type     = r["input_type"]
                st.session_state.lin_downstream     = r["downstream"]
                st.session_state.lin_error          = None
                st.session_state.lin_results        = None
                st.session_state.lin_chain_analysis = None
            else:
                st.session_state.lin_results    = results
                st.session_state.lin_tree       = None
                st.session_state.lin_error      = None

    # ── Render error
    if st.session_state.lin_error:
        st.error(f"❌ {st.session_state.lin_error}")

    # ── Render batch results
    elif st.session_state.get("lin_results"):
        results = st.session_state["lin_results"]
        st.markdown(f"#### 📋 Batch — {len(results)} quer{'y' if len(results)==1 else 'ies'}")
        rows_tbl = []
        for r in results:
            t,u,i,score,cx = compute_complexity(r["stats"])
            rows_tbl.append({"Query":r["query"],"Provider":r["provider"],"Type":r["input_type"],
                             "Complexity":cx,"Score":score,"Transforms":t,"Update Rules":u,"InfoPkgs":i})
        st.dataframe(pd.DataFrame(rows_tbl), use_container_width=True)
        # Single combined Excel with all queries + downstream + all BEx names
        combined_buf = batch_to_excel(results)
        dl1, dl2 = st.columns(2)
        with dl1:
            st.download_button(
                f"⬇️ Download All {len(results)} Lineages (.xlsx)",
                data=combined_buf,
                file_name="bw_lineage_all.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="batch_lin_dl",
                use_container_width=True,
            )
        with dl2:
            # ── Combined ZIP — Excel reports and TXT reports kept in their own
            # subfolders inside one archive, so a single click gets everything.
            zip_all_buf = io.BytesIO()
            with zipfile.ZipFile(zip_all_buf, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.writestr("excel/bw_lineage_all.xlsx", combined_buf.getvalue())
                for r in results:
                    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in r["query"])
                    xbuf = tree_to_excel(r["query"], r["provider"], r["tree"],
                                         downstream=r.get("downstream"),
                                         input_type=r.get("input_type","QUERY"))
                    zf.writestr(f"excel/lineage_{safe}.xlsx", xbuf.getvalue())
                    tbuf = tree_to_txt_download(r["query"], r["provider"], r["tree"],
                                                r["stats"], downstream=r.get("downstream"),
                                                input_type=r.get("input_type","QUERY"))
                    zf.writestr(f"txt/lineage_{safe}.txt", tbuf)
            zip_all_buf.seek(0)
            st.download_button(
                f"📦 Download All {len(results)} (Excel + TXT .zip)",
                data=zip_all_buf,
                file_name="bw_lineage_all_reports.zip",
                mime="application/zip",
                key="batch_zip_dl",
                use_container_width=True,
            )
        for r in results:
            t,u,i,score,cx = compute_complexity(r["stats"])
            with st.expander(f"🔍 {r['query']}  —  {cx}  (score {score})", expanded=False):
                mc1,mc2,mc3,mc4 = st.columns(4)
                mc1.metric("🔁 Transforms",t); mc2.metric("🔄 Update Rules",u)
                mc3.metric("📬 InfoPkgs",f"{i:,}"); mc4.metric("🎯 Score",score)
                render_lineage_visual(flatten_tree(r["tree"]))
                if r.get("downstream"):
                    render_downstream_visual(r["downstream"], dfs=dfs)
                safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in r["query"])
                xbuf = tree_to_excel(r["query"], r["provider"], r["tree"],
                                     downstream=r.get("downstream"),
                                     input_type=r.get("input_type","QUERY"))
                bc1, bc2 = st.columns(2)
                with bc1:
                    st.download_button("⬇️ Excel", data=xbuf,
                                       file_name=f"lineage_{safe}.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       key=f"dl_{safe}", use_container_width=True)
                with bc2:
                    tbuf = tree_to_txt_download(r["query"], r["provider"], r["tree"],
                                                r["stats"], downstream=r.get("downstream"),
                                                input_type=r.get("input_type","QUERY"))
                    st.download_button("⬇️ TXT", data=tbuf,
                                       file_name=f"lineage_{safe}.txt",
                                       mime="text/plain",
                                       key=f"dl_txt_{safe}", use_container_width=True)

    # ── Render single result
    elif st.session_state.lin_tree is not None:
        tree       = st.session_state.lin_tree
        stats      = st.session_state.lin_stats
        lq         = st.session_state.lin_query
        prov       = str(st.session_state.get("lin_provider",""))
        input_type = st.session_state.get("lin_input_type","QUERY")
        downstream = st.session_state.get("lin_downstream")
        t, u, i, score, cx_lin = compute_complexity(stats)

        if input_type == "DATASOURCE":
            st.markdown(f"**Bidirectional Lineage for** `{lq}`")
            st.info("📡 DataSource/Extractor detected — showing full downstream chain (InfoPackage → DataSource → Transformations → InfoProvider → BEx Queries)")
        else:
            st.markdown(f"**Lineage for** `{lq}`")

        mc1,mc2,mc3,mc4 = st.columns(4)

        if input_type == "DATASOURCE" and downstream:
            ds_transforms = len(set(
                ch["transformation_up"] for ch in downstream if ch.get("transformation_up")
            ) | set(
                ch["transformation_dn"] for ch in downstream if ch.get("transformation_dn")
            ))
            ds_ips  = len(set(
                ip for ch in downstream for ip in ch.get("infopackages", [])
            ))
            ds_queries = len(set(
                q for ch in downstream for q in ch.get("queries", [])
            ))
            mc1.metric("⚙️ Transforms",  ds_transforms)
            mc2.metric("🔄 Update Rules", u)
            mc3.metric("📬 InfoPkgs",     ds_ips)
            mc4.metric("🔍 BEx Queries",  ds_queries)
        else:
            mc1.metric("🔁 Transforms",  t)
            mc2.metric("🔄 Update Rules", u)
            mc3.metric("📬 InfoPkgs",    f"{i:,}")
            mc4.metric("🎯 Score",        score)
        st.markdown("")

        if input_type == "DATASOURCE":
            # Show downstream only — InfoPackages are already shown inside downstream chains
            if downstream:
                render_downstream_visual(downstream, dfs=dfs)
            else:
                st.caption("No downstream transformation chains found from this DataSource.")
        else:
            # BEx Query — show lineage + AI Transformation Analysis tabs
            tab_vis, tab_ai = st.tabs(["🔗 Lineage", "🤖 AI Transformation Analysis"])

            with tab_vis:
                render_lineage_visual(flatten_tree(tree))

            with tab_ai:
                tranids = list(stats.get("transformations", set()))
                if not tranids:
                    st.info("No transformations found in this lineage chain.")
                else:
                    st.caption(
                        f"**{len(tranids)} transformation(s)** found in this lineage chain. "
                        "AI analyses each one and provides an overall migration assessment."
                    )
                    chain_res = st.session_state.get("lin_chain_analysis")

                    if chain_res is None:
                        if st.button("🤖 Analyse All Transformations", type="primary", key="run_chain"):
                            with st.spinner(f"Analysing {len(tranids)} transformation(s)…"):
                                st.session_state.lin_chain_analysis = do_chain_analysis(
                                    tranids, dfs
                                )
                            st.rerun()
                    else:
                        # ── Overall narrative
                        overall = chain_res.get("_overall","")
                        if overall and not overall.startswith("[ERROR"):
                            st.markdown(
                                f'<div class="ai-sum-box">🔗 <b>Chain Overview</b>'
                                f'<br><br>{overall}</div>',
                                unsafe_allow_html=True
                            )
                        elif overall.startswith("[ERROR"):
                            st.error(overall)

                        st.markdown("")
                        st.markdown(f"**Individual transformations ({len(tranids)}):**")

                        # ── Per-transformation cards
                        trfn_df = dfs.get("trfn_df")
                        for tid in sorted(tranids):
                            info = chain_res.get(tid, {})
                            src  = info.get("sourcename", "—")
                            tgt  = info.get("targetname", "—")
                            with st.expander(f"⚙️ {tid}", expanded=False):
                                c1, c2 = st.columns(2)
                                c1.markdown(f"**Source:** `{src}`")
                                c2.markdown(f"**Target:** `{tgt}`")
                                st.markdown("")

                                # Per-transformation AI on demand
                                ai_key = f"trfn_ai_{tid}"
                                if st.session_state.get(ai_key) is None:
                                    if st.button("✨ Explain this transformation",
                                                 key=f"btn_{tid}"):
                                        with st.spinner("Analysing…"):
                                            st.session_state[ai_key] = do_single_trfn_ai(
                                                tid, src, tgt,
                                                info.get("sourcetype",""),
                                                info.get("targettype","")
                                            )
                                        st.rerun()
                                else:
                                    ai_txt = st.session_state[ai_key]
                                    if ai_txt.startswith("[ERROR"):
                                        st.error(ai_txt)
                                    else:
                                        st.markdown(
                                            f'<div class="ai-sum-box">{ai_txt}</div>',
                                            unsafe_allow_html=True
                                        )
                                    if st.button("🔄 Regenerate", key=f"regen_{tid}"):
                                        st.session_state[ai_key] = None; st.rerun()

                        st.markdown("")
                        if st.button("🔄 Re-analyse all", key="regen_chain"):
                            st.session_state.lin_chain_analysis = None
                            # Clear individual transformation AI results
                            for k in list(st.session_state.keys()):
                                if k.startswith("trfn_ai_"):
                                    del st.session_state[k]
                            st.rerun()

        st.markdown("")
        dl1, dl2 = st.columns(2)
        excel_buf = tree_to_excel(lq, prov, tree,
                                   downstream=downstream,
                                   input_type=input_type)
        lbl = "Bidirectional" if input_type == "DATASOURCE" else "Lineage"
        with dl1:
            st.download_button(f"⬇️ Download {lbl} (.xlsx)", data=excel_buf,
                               file_name=f"lineage_{lq}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="lin_dl", use_container_width=True)
        with dl2:
            txt_buf = tree_to_txt_download(lq, prov, tree, stats,
                                           downstream=downstream,
                                           input_type=input_type)
            st.download_button(f"⬇️ Download {lbl} (.txt)", data=txt_buf,
                               file_name=f"lineage_{lq}.txt",
                               mime="text/plain",
                               key="lin_dl_txt", use_container_width=True)



# =====================================================
# FOOTER
# =====================================================
st.markdown("---")
st.caption("Gyansys Made AI BW Migration Toolkit — © 2026 Gyansys. All rights reserved.")